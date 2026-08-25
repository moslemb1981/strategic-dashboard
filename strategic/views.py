import logging
import re
import io
import math
import json
import requests

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, Http404, HttpResponse
from django.urls import reverse
from django.db.models import Q

from .jalali_utils import jalali_str_to_gregorian, gregorian_to_jalali_str

from .models import (
    Study, Initiative, Risk, SWOTItem, TOWSStrategy, StrategicObjective, Competitor, PestelFactor,
    BusinessUnit, StrategyTheme, PorterForce, OrgIdentity, OrgValue, QualityPolicyPoint, McKinsey7S,
    ValueChainActivity, Stakeholder, CrossImpactFactor, CrossImpactLink, Scenario, ScenarioAxes,
    CompanyObjective, CompanyKPI, Document, StrategicKPI, LegalRequirement, EnvironmentalFactor,
    ScenarioHighlight, OperationalKPI, ScenarioResponseStrategy, RawIdentifiedFactor, MarketIntelRecord,
)
from .forms import (
    StudyForm, InitiativeForm, RiskForm, SWOTItemForm, TOWSStrategyForm, StrategicObjectiveForm,
    CompetitorForm, PestelFactorForm, StrategyThemeForm, PorterForceForm, McKinsey7SForm, ValueChainActivityForm,
    StakeholderForm, CrossImpactFactorForm, ScenarioForm, ScenarioAxesForm, CompanyObjectiveForm, CompanyKPIForm, DocumentForm,
    StrategicKPIForm, LegalRequirementForm, EnvironmentalFactorForm, OperationalKPIForm, ScenarioResponseStrategyForm,
    clean_number_string,
)

logger = logging.getLogger("strategic")


def _has_perm(request, perm):
    """Checks a Django model permission (e.g. 'strategic.add_study').
    Superusers always pass. On failure, flashes a message the user sees."""
    if request.user.has_perm(perm):
        return True
    messages.error(request, "شما اجازه انجام این عملیات را ندارید. برای دسترسی ویرایش با مدیر سیستم هماهنگ کنید.")
    logger.warning("PERMISSION DENIED: user=%s perm=%s", request.user, perm)
    return False


def _log_action(request, action, label):
    logger.info("%s: user=%s item=%r", action, request.user, label)


def _swot_code_map():
    """کد S1/W1/O1/T1 هیچ‌جا توی دیتابیس ذخیره نمی‌شه — دقیقاً همون‌طور که توی خود صفحه‌ی
    SWOT لحظه‌ای ساخته می‌شه (ترتیب هر دسته داخل هر کسب‌وکار، بر اساس -weight, created_at)،
    این تابع عیناً همون منطق رو تکرار می‌کنه تا کدها همه‌جای سامانه با صفحه‌ی SWOT یکی باشن."""
    code_map, counters = {}, {}
    items = SWOTItem.objects.all().order_by("business_unit_id", "category", "-weight", "created_at")
    for si in items:
        key = (si.business_unit_id, si.category)
        counters[key] = counters.get(key, 0) + 1
        code_map[si.pk] = f"{si.category.upper()}{counters[key]}"
    return code_map


def home(request):
    objectives = list(StrategicObjective.objects.all())
    obj_total = len(objectives)
    obj_score = 0
    if obj_total:
        weight = {"on": 1, "watch": 0.5, "risk": 0}
        obj_score = round(sum(weight.get(o.status, 0) for o in objectives) / obj_total * 100)

    initiatives = list(Initiative.objects.all())
    init_total = len(initiatives)
    init_behind = sum(1 for i in initiatives if i.status == "needs_attention")
    init_on_track = init_total - init_behind
    init_avg_progress = round(sum(i.progress for i in initiatives) / init_total) if init_total else 0

    studies = list(Study.objects.all())
    study_total = len(studies)
    study_done = sum(1 for s in studies if s.status == "done")
    study_pct = round(study_done / study_total * 100) if study_total else 0

    risks = list(Risk.objects.all())
    risk_total = len(risks)
    risk_high = sum(1 for r in risks if r.zone in ("high", "crit"))
    risk_pct = round(risk_high / risk_total * 100) if risk_total else 0

    # فید فعالیت‌های اخیر — از هر ۷ مدل، آخرین رکوردها را ترکیب می‌کند
    # برای کاربر بدون‌لاگین، موارد ریسک و SWOT (حساس) از فید حذف می‌شوند
    activity = []
    for s in Study.objects.order_by("-created_at")[:5]:
        activity.append({"icon": "fa-book", "text": f"مطالعه «{s.title}» ثبت شد", "tag": "کتابخانه مطالعات", "dt": s.created_at})
    for i in Initiative.objects.order_by("-created_at")[:5]:
        activity.append({"icon": "fa-route", "text": f"ابتکار «{i.title}» ثبت شد", "tag": "نقشه راه", "dt": i.created_at})
    if request.user.is_authenticated:
        for r in Risk.objects.order_by("-created_at")[:5]:
            activity.append({"icon": "fa-triangle-exclamation", "text": f"ریسک «{r.title}» ثبت شد", "tag": "نقشه ریسک", "dt": r.created_at})
    for o in StrategicObjective.objects.order_by("-created_at")[:5]:
        activity.append({"icon": "fa-map", "text": f"هدف «{o.code} — {o.title}» ثبت شد", "tag": "نقشه استراتژیک", "dt": o.created_at})
    if request.user.is_authenticated:
        for it in SWOTItem.objects.order_by("-created_at")[:5]:
            activity.append({"icon": "fa-table-cells", "text": f"مورد SWOT «{it.text}» ثبت شد", "tag": "SWOT", "dt": it.created_at})
    for c in Competitor.objects.order_by("-created_at")[:5]:
        activity.append({"icon": "fa-chart-line", "text": f"بازیگر «{c.name}» ثبت شد", "tag": "هوش رقابتی", "dt": c.created_at})
    for f in PestelFactor.objects.order_by("-created_at")[:5]:
        activity.append({"icon": "fa-earth-americas", "text": f"عامل «{f.text}» ثبت شد", "tag": "PESTEL", "dt": f.created_at})

    activity.sort(key=lambda a: a["dt"], reverse=True)
    activity = activity[:6]

    swot_count = SWOTItem.objects.count()
    org_value_count = OrgValue.objects.count()
    quality_policy_count = QualityPolicyPoint.objects.count()
    org_identity = OrgIdentity.objects.first()
    vision_count = 1 if (org_identity and org_identity.vision) else 0
    mission_count = 1 if (org_identity and org_identity.mission) else 0

    home_kpis_qs = [
        k for k in CompanyKPI.objects.exclude(target_1405="").exclude(actual_1405="")
        if k.progress_pct_1405 is not None
    ]
    home_kpis_total = len(home_kpis_qs)
    home_kpis = home_kpis_qs[:8]
    scenario_total = Scenario.objects.count()
    pestel_count = PestelFactor.objects.count()
    cross_impact_count = CrossImpactFactor.objects.count()
    company_objective_count = CompanyObjective.objects.count()
    company_kpi_count = CompanyKPI.objects.count()
    legal_requirement_count = LegalRequirement.objects.count()
    environmental_factor_count = EnvironmentalFactor.objects.count()
    stakeholder_count = Stakeholder.objects.count()
    porter_count = PorterForce.objects.count()
    mckinsey7s_count = McKinsey7S.objects.count()
    value_chain_count = ValueChainActivity.objects.count()
    tows_count = TOWSStrategy.objects.count()
    operational_kpi_count = OperationalKPI.objects.count()

    swot_breakdown = {c: SWOTItem.objects.filter(category=c).count() for c in ["s", "w", "o", "t"]}
    persp_breakdown = {p: StrategicObjective.objects.filter(perspective=p).count() for p in ["financial", "customer", "process", "learning"]}

    obj_status_breakdown = {s: sum(1 for o in objectives if o.status == s) for s in ["on", "watch", "risk"]}
    obj_status_pct = {
        k: (round(v / obj_total * 100) if obj_total else 0) for k, v in obj_status_breakdown.items()
    }
    init_status_breakdown = {
        "running": sum(1 for i in initiatives if i.status == "in_progress"),
        "planned": sum(1 for i in initiatives if i.status == "deviation"),
        "done": sum(1 for i in initiatives if i.status == "done"),
    }
    init_status_pct = {
        k: (round(v / init_total * 100) if init_total else 0) for k, v in init_status_breakdown.items()
    }

    total_records = (
        legal_requirement_count + stakeholder_count + environmental_factor_count + pestel_count + porter_count
        + mckinsey7s_count + value_chain_count + cross_impact_count + scenario_total + swot_count
        + org_value_count + quality_policy_count + obj_total + init_total + risk_total
        + company_objective_count + company_kpi_count
    )

    return render(request, "strategic/home.html", {
        "active_page": "home",
        "obj_score": obj_score, "obj_total": obj_total,
        "obj_status_breakdown": obj_status_breakdown, "obj_status_pct": obj_status_pct,
        "init_total": init_total, "init_on_track": init_on_track, "init_behind": init_behind, "init_avg_progress": init_avg_progress,
        "init_status_breakdown": init_status_breakdown, "init_status_pct": init_status_pct,
        "study_total": study_total, "study_done": study_done, "study_pct": study_pct,
        "risk_total": risk_total, "risk_high": risk_high, "risk_pct": risk_pct,
        "competitor_count": Competitor.objects.count(),
        "pestel_count": pestel_count,
        "cross_impact_count": cross_impact_count,
        "scenario_selected": Scenario.objects.filter(is_selected=True).first(),
        "company_objective_count": company_objective_count,
        "company_kpi_count": company_kpi_count,
        "operational_kpi_count": operational_kpi_count,
        "document_count": Document.objects.count(),
        "legal_requirement_count": legal_requirement_count,
        "environmental_factor_count": environmental_factor_count,
        "stakeholder_count": stakeholder_count,
        "porter_count": porter_count,
        "mckinsey7s_count": mckinsey7s_count,
        "value_chain_count": value_chain_count,
        "swot_count": swot_count,
        "swot_breakdown": swot_breakdown,
        "tows_count": tows_count,
        "persp_breakdown": persp_breakdown,
        "org_value_count": org_value_count,
        "quality_policy_count": quality_policy_count,
        "vision_count": vision_count,
        "mission_count": mission_count,
        "home_kpis": home_kpis,
        "home_kpis_total": home_kpis_total,
        "scenario_total": scenario_total,
        "total_records": total_records,
        "activity": activity,
    })


# ---------------- Research library ----------------

def research(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_study" if obj_id else "strategic.add_study"
        if _has_perm(request, perm):
            instance = get_object_or_404(Study, pk=obj_id) if obj_id else None
            form = StudyForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE Study" if obj_id else "CREATE Study", str(form.instance))
                return redirect("strategic:research")
        else:
            form = StudyForm()
    else:
        form = StudyForm()

    studies = Study.objects.all()
    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    if q:
        studies = studies.filter(title__icontains=q)
    if status:
        studies = studies.filter(status=status)

    return render(request, "strategic/research.html", {
        "active_page": "research", "studies": studies, "form": form, "q": q, "status": status,
    })


# ---------------- ذینفعان ----------------

def stakeholders(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_stakeholder" if obj_id else "strategic.add_stakeholder"
        if _has_perm(request, perm):
            instance = get_object_or_404(Stakeholder, pk=obj_id) if obj_id else None
            form = StakeholderForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE Stakeholder" if obj_id else "CREATE Stakeholder", str(form.instance))
                return redirect("strategic:stakeholders")
        else:
            form = StakeholderForm()
    else:
        form = StakeholderForm()

    items = Stakeholder.objects.all().prefetch_related("related_porters", "linked_pestel_factors", "swot_items__business_unit")
    q = request.GET.get("q", "").strip()
    dept = request.GET.get("dept", "").strip()
    if q:
        items = items.filter(
            Q(name__icontains=q) | Q(need__icontains=q) |
            Q(risk_text__icontains=q) | Q(opportunity_text__icontains=q)
        )
    if dept:
        items = items.filter(department=dept)

    all_items = list(Stakeholder.objects.all().prefetch_related("related_initiatives"))

    risk_scores_sorted = sorted((i.risk_score for i in all_items if i.risk_score), reverse=True)
    n_risk = max(1, round(len(risk_scores_sorted) * 0.2)) if risk_scores_sorted else 0
    top_risks = sorted(
        [i for i in all_items if i.risk_score],
        key=lambda i: i.risk_score, reverse=True,
    )[:n_risk]

    opp_scores_sorted = sorted((i.opportunity_score for i in all_items if i.opportunity_score), reverse=True)
    n_opp = max(1, round(len(opp_scores_sorted) * 0.2)) if opp_scores_sorted else 0
    top_opportunities = sorted(
        [i for i in all_items if i.opportunity_score],
        key=lambda i: i.opportunity_score, reverse=True,
    )[:n_opp]
    departments = sorted({i.department for i in all_items if i.department})

    total = len(all_items)
    internal_n = sum(1 for i in all_items if i.is_internal)
    external_n = sum(1 for i in all_items if i.is_external)
    with_need_n = sum(1 for i in all_items if i.need)
    with_risk_n = sum(1 for i in all_items if i.risk_text)
    with_opp_n = sum(1 for i in all_items if i.opportunity_text)
    critical_with_action_n = sum(
        1 for i in top_risks if i.action or i.related_initiatives.exists()
    )

    def pct(part, whole):
        return round(part / whole * 100) if whole else 0

    ie_sum = max(internal_n + external_n, 1)
    summary = {
        "total": total,
        "internal": internal_n, "external": external_n,
        "internal_pct": round(internal_n / ie_sum * 100), "external_pct": round(external_n / ie_sum * 100),
        "with_need": with_need_n, "with_need_pct": pct(with_need_n, total),
        "with_risk": with_risk_n, "with_risk_pct": pct(with_risk_n, total),
        "with_opportunity": with_opp_n, "with_opportunity_pct": pct(with_opp_n, total),
        "critical_total": len(top_risks),
        "critical_with_action": critical_with_action_n,
        "critical_with_action_pct": pct(critical_with_action_n, len(top_risks)),
    }

    return render(request, "strategic/stakeholders.html", {
        "active_page": "stakeholders", "items": items, "form": form, "q": q, "dept": dept,
        "departments": departments, "top_risks": top_risks, "top_opportunities": top_opportunities,
        "porter_forces": PorterForce.objects.all(), "summary": summary,
        "total_count": len(all_items),
    })


@login_required
def stakeholder_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_stakeholder"):
        obj = get_object_or_404(Stakeholder, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE Stakeholder", label)
    return redirect("strategic:stakeholders")


@login_required
def study_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_study"):
        _obj = get_object_or_404(Study, pk=pk)
        _label = str(_obj)
        _obj.delete()
        _log_action(request, "DELETE Study", _label)
    return redirect("strategic:research")


# ---------------- Roadmap / initiatives ----------------

def roadmap(request):
    business_units = list(BusinessUnit.objects.all())
    group_mode = request.POST.get("group_mode") or request.GET.get("group_mode") or "bu"
    if group_mode not in ("bu", "division", "work_group"):
        group_mode = "bu"

    bu_id = request.POST.get("business_unit") or request.GET.get("bu")
    current_bu = None
    if bu_id:
        current_bu = next((b for b in business_units if str(b.pk) == str(bu_id)), None)
    if not current_bu and business_units:
        current_bu = business_units[0]

    group_field = {"division": "division", "work_group": "work_group"}.get(group_mode)
    group_tabs, current_group_key = [], None
    if group_field:
        distinct_values = sorted(set(
            v.strip() for v in Initiative.objects.exclude(**{group_field: ""}).values_list(group_field, flat=True) if v and v.strip()
        ))
        counts = {}
        for v in distinct_values:
            counts[v] = Initiative.objects.filter(**{group_field: v}).count()
        group_tabs = [{"key": v, "count": counts[v]} for v in distinct_values]
        current_group_key = request.POST.get("g") or request.GET.get("g") or (distinct_values[0] if distinct_values else None)

    bu_counts = {b.pk: Initiative.objects.filter(business_unit=b).count() for b in business_units}
    for b in business_units:
        b.init_count = bu_counts.get(b.pk, 0)

    type_filter = request.POST.get("item_type") or request.GET.get("item_type") or "all"
    if type_filter not in ("all", "project", "action"):
        type_filter = "all"

    view_mode = request.POST.get("view") or request.GET.get("view") or "card"
    if view_mode not in ("card", "list"):
        view_mode = "card"

    if request.method == "POST" and request.POST.get("form_kind", "initiative") == "initiative":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_initiative" if obj_id else "strategic.add_initiative"
        if _has_perm(request, perm):
            instance = get_object_or_404(Initiative, pk=obj_id) if obj_id else None
            form = InitiativeForm(request.POST, instance=instance, business_unit=current_bu)
            if form.is_valid():
                obj = form.save(commit=False)
                if not obj_id and current_bu:
                    obj.business_unit = current_bu
                obj.save()
                form.save_m2m()
                _log_action(request, "UPDATE Initiative" if obj_id else "CREATE Initiative", str(obj))
                params = f"?group_mode={group_mode}"
                params += f"&g={current_group_key}" if group_field and current_group_key else f"&bu={current_bu.pk}" if current_bu else ""
                return redirect(reverse("strategic:roadmap") + params)
        else:
            form = InitiativeForm(business_unit=current_bu)
    else:
        form = InitiativeForm(business_unit=current_bu)

    base_prefetch = ("objectives", "source_kpi", "source_operational_kpi", "source_tows", "source_risk",
                      "source_tows__source_items", "linked_stakeholders")
    if group_field:
        initiatives_qs = (
            Initiative.objects.filter(**{group_field: current_group_key})
            .prefetch_related(*base_prefetch)
            if current_group_key else Initiative.objects.none()
        )
    else:
        initiatives_qs = (
            Initiative.objects.filter(business_unit=current_bu).prefetch_related(*base_prefetch)
            if current_bu else Initiative.objects.none()
        )
    if type_filter != "all":
        initiatives_qs = initiatives_qs.filter(item_type=type_filter)
    # پروژه‌ها همیشه قبل از اقدام‌ها نمایش داده می‌شوند (سپس بر اساس تاریخ شروع)
    initiatives = sorted(initiatives_qs, key=lambda i: (0 if i.item_type == "project" else 1, i.start_date))

    STATUS_ORDER = [
        ("deviation", "انحراف پروژه", True),
        ("in_progress", "در حال اجرا", True),
        ("done", "تکمیل‌شده", False),
    ]
    status_groups = []
    for key, label, open_default in STATUS_ORDER:
        items = [i for i in initiatives if i.status == key]
        status_groups.append({
            "key": key, "label": label, "items": items, "count": len(items), "open_default": open_default,
        })

    return render(request, "strategic/roadmap.html", {
        "active_page": "roadmap", "initiatives": initiatives, "status_groups": status_groups, "form": form,
        "business_units": business_units, "current_bu": current_bu,
        "group_mode": group_mode, "group_tabs": group_tabs, "current_group_key": current_group_key, "bu_counts": bu_counts,
        "type_filter": type_filter, "view_mode": view_mode,
    })


_INITIATIVE_EXCEL_HEADERS = [
    "کد پروژه", "عنوان پروژه", "نوع (پروژه/اقدام)", "واحد مسئول", "کارگروه", "معاونت", "کسب‌وکار", "تاریخ شروع (شمسی)",
    "تاریخ پایان (شمسی)", "پیشرفت (٪)", "وضعیت",
]
_INITIATIVE_STATUS_LABEL_TO_KEY = {label: key for key, label in Initiative.STATUS_CHOICES}
_INITIATIVE_TYPE_LABEL_TO_KEY = {label: key for key, label in Initiative.TYPE_CHOICES}


@login_required
@login_required
def bsc_sync(request):
    if request.method != "POST" or not request.user.is_superuser:
        messages.error(request, "این عملیات مجاز نیست.")
        return redirect("strategic:roadmap")

    try:
        import pyodbc
    except ImportError:
        return render(request, "strategic/bsc_sync_report.html", {
            "connection_error": "کتابخانه‌ی pyodbc روی سرور نصب نیست.",
        })

    try:
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=localhost\\SQLEXPRESS;DATABASE=BSC;Trusted_Connection=yes;",
            timeout=10,
        )
    except Exception as e:
        return render(request, "strategic/bsc_sync_report.html", {
            "connection_error": str(e),
        })

    try:
        cursor = conn.cursor()
        updated, not_found = [], []
        initiatives = Initiative.objects.exclude(code__isnull=True).exclude(code="")

        for init in initiatives:
            cursor.execute("SELECT id FROM projects WHERE code = ?", init.code)
            row = cursor.fetchone()
            if not row:
                not_found.append({"code": init.code, "title": init.title})
                continue

            project_id = row[0]
            cursor.execute(
                "SELECT TOP 1 calced_val FROM project_values "
                "WHERE project_id = ? AND calced_val IS NOT NULL "
                "ORDER BY period_id DESC",
                project_id,
            )
            prow = cursor.fetchone()
            if not prow or prow[0] is None:
                not_found.append({"code": init.code, "title": init.title})
                continue

            new_progress = max(0, min(100, round(prow[0])))
            if init.progress != new_progress:
                init.progress = new_progress
                init.save(update_fields=["progress"])
            updated.append({"code": init.code, "title": init.title, "progress": new_progress})
    finally:
        conn.close()

    _log_action(request, "SYNC progress from BSC",
                f"{len(updated)} به‌روزشده، {len(not_found)} یافت‌نشده")

    return render(request, "strategic/bsc_sync_report.html", {
        "updated": updated,
        "not_found": not_found,
        "updated_count": len(updated),
        "not_found_count": len(not_found),
    })


def initiative_export(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:roadmap")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "پروژه‌های تحول"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(_INITIATIVE_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_i, i in enumerate(Initiative.objects.all().select_related("business_unit"), start=2):
        values = [
            i.code, i.title, i.get_item_type_display(), i.owner, i.work_group, i.division,
            i.business_unit.name if i.business_unit else "",
            gregorian_to_jalali_str(i.start_date) if i.start_date else "",
            gregorian_to_jalali_str(i.end_date) if i.end_date else "",
            i.progress, i.get_status_display(),
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_i, column=col, value=val)

    widths = [14, 34, 16, 20, 22, 22, 22, 14, 14, 10, 16]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT Initiative Excel", f"{Initiative.objects.count()} ردیف")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="projeh-haye-tahavol.xlsx"'
    return response


@login_required
def initiative_import(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:roadmap")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "فایلی انتخاب نشده است.")
        return redirect("strategic:roadmap")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"], data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "فایل اکسل قابل خواندن نیست. لطفاً فرمت را بررسی کنید.")
        return redirect("strategic:roadmap")

    def _s(v):
        return "" if v is None else str(v).strip()

    business_units_by_name = {b.name: b for b in BusinessUnit.objects.all()}
    created, updated, skipped = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        code = _s(row[0]) if len(row) > 0 else ""
        title = _s(row[1])
        if not title:
            skipped += 1
            continue
        owner = _s(row[3]) if len(row) > 3 else ""
        work_group = _s(row[4]) if len(row) > 4 else ""
        division = _s(row[5]) if len(row) > 5 else ""
        bu_name = _s(row[6]) if len(row) > 6 else ""
        business_unit = business_units_by_name.get(bu_name)
        try:
            start_date = jalali_str_to_gregorian(_s(row[7])) if len(row) > 7 and _s(row[7]) else None
            end_date = jalali_str_to_gregorian(_s(row[8])) if len(row) > 8 and _s(row[8]) else None
        except Exception:
            start_date, end_date = None, None
        if not start_date or not end_date:
            skipped += 1
            continue
        try:
            progress = int(row[9]) if len(row) > 9 and row[9] not in (None, "") else 0
        except (TypeError, ValueError):
            progress = 0
        status_label = _s(row[10]) if len(row) > 10 else ""
        status = _INITIATIVE_STATUS_LABEL_TO_KEY.get(status_label, "in_progress")
        type_label = _s(row[2]) if len(row) > 2 else ""
        item_type = _INITIATIVE_TYPE_LABEL_TO_KEY.get(type_label, "project")

        if code:
            existing = Initiative.objects.filter(code=code).first()
        else:
            existing = Initiative.objects.filter(title=title, code="").first()
        defaults = dict(
            code=code, item_type=item_type, owner=owner, work_group=work_group, division=division, business_unit=business_unit,
            start_date=start_date, end_date=end_date, progress=progress, status=status,
        )
        if existing:
            for k, v in defaults.items():
                setattr(existing, k, v)
            existing.save()
            updated += 1
        else:
            Initiative.objects.create(title=title, **defaults)
            created += 1

    _log_action(request, "IMPORT Initiative Excel", f"{created} جدید، {updated} به‌روزشده، {skipped} رد‌شده")
    messages.success(request, f"وارد کردن انجام شد: {created} پروژه جدید، {updated} پروژه به‌روزرسانی‌شده، {skipped} ردیف نامعتبر رد شد.")
    return redirect("strategic:roadmap")


@login_required
def initiative_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_initiative"):
        _obj = get_object_or_404(Initiative, pk=pk)
        _label = str(_obj)
        bu_pk = _obj.business_unit_id
        _obj.delete()
        _log_action(request, "DELETE Initiative", _label)
        if bu_pk:
            return redirect(reverse("strategic:roadmap") + f"?bu={bu_pk}")
    return redirect("strategic:roadmap")


# ---------------- Market / competitive intelligence ----------------

def market_intel(request):
    records = list(MarketIntelRecord.objects.all()[:200])
    grouped = {}
    for r in records:
        grouped.setdefault(r.category, []).append(r)
    category_labels = dict(MarketIntelRecord.CATEGORY_CHOICES)
    latest_fetch = records[0].fetched_at if records else None

    return render(request, "strategic/market_intel.html", {
        "active_page": "market_intel", "grouped": grouped, "category_labels": category_labels,
        "categories": MarketIntelRecord.CATEGORY_CHOICES, "latest_fetch": latest_fetch,
        "total_count": len(records),
    })


@login_required
def market_intel_fetch(request):
    if request.method != "POST" or not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم و از طریق دکمه‌ی مربوطه مجاز است.")
        return redirect("strategic:market_intel")

    created, errors = [], []

    # --- نمونه‌ی اول: نرخ برابری چند ارز اصلی (منبع: Frankfurter/ECB، عمومی و بدون نیاز به کلید) ---
    # ⚠️ این یک نمونه‌ی عمومی است، نه نرخ ریال ایران — چون منابع استاندارد بین‌المللی نرخ ریال را
    # به‌دلیل تحریم‌ها ندارند. برای نرخ دلار/ریال باید یک منبع اختصاصی ایرانی انتخاب و بعداً اضافه شود.
    try:
        resp = requests.get("https://api.frankfurter.app/latest?from=USD&to=EUR,GBP", timeout=8)
        resp.raise_for_status()
        data = resp.json()
        rate_date = data.get("date", "")
        for currency, rate in data.get("rates", {}).items():
            rec = MarketIntelRecord.objects.create(
                category="exchange_rate",
                title=f"نرخ برابری USD به {currency}",
                value=str(rate),
                unit=currency,
                source_name="Frankfurter (ECB)",
                source_url="https://www.frankfurter.app/",
                fetched_by=request.user,
            )
            created.append(rec.title)
    except Exception as e:
        errors.append(f"نرخ ارز: {e}")

    # --- نمونه‌ی دوم: نرخ دلار به ریال (منبع: TGJU.org — چون منابع بین‌المللی استاندارد،
    # نرخ ریال ایران رو به‌خاطر تحریم‌ها ندارن) ---
    # روش کار: توی هر صفحه‌ی TGJU، یه نوار «قیمت لحظه‌ای» هست که شامل «دلار NUMBER (درصد%)»
    # می‌شه — این الگو رو با regex پیدا می‌کنیم، نه صفحه‌بندی خاصی که ممکنه عوض بشه.
    try:
        resp = requests.get(
            "https://www.tgju.org/profile/price_dollar_rl",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
            timeout=10,
        )
        resp.raise_for_status()
        m = re.search(r"دلار[^\d]{0,40}([\d,]{6,12})", resp.text)
        if not m:
            raise ValueError("الگوی قیمت دلار توی صفحه پیدا نشد (شاید ساختار سایت عوض شده)")
        rial_value = m.group(1).replace(",", "")
        rec = MarketIntelRecord.objects.create(
            category="exchange_rate",
            title="نرخ دلار آزاد به ریال",
            value=rial_value,
            unit="ریال",
            source_name="TGJU.org",
            source_url="https://www.tgju.org/profile/price_dollar_rl",
            fetched_by=request.user,
        )
        created.append(rec.title)
    except Exception as e:
        errors.append(f"نرخ دلار (TGJU): {e}")

    if created:
        _log_action(request, "FETCH MarketIntel", f"{len(created)} رکورد: {', '.join(created)}")
        messages.success(request, f"به‌روزرسانی انجام شد: {len(created)} رکورد جدید ثبت شد.")
    if errors:
        messages.error(request, "برخی موارد با خطا مواجه شدند: " + " | ".join(errors))
    if not created and not errors:
        messages.warning(request, "هیچ داده‌ی جدیدی دریافت نشد.")

    return redirect("strategic:market_intel")


def market(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_competitor" if obj_id else "strategic.add_competitor"
        if _has_perm(request, perm):
            instance = get_object_or_404(Competitor, pk=obj_id) if obj_id else None
            form = CompetitorForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE Competitor" if obj_id else "CREATE Competitor", str(form.instance))
                return redirect("strategic:market")
        else:
            form = CompetitorForm()
    else:
        form = CompetitorForm()

    competitors = Competitor.objects.all()
    return render(request, "strategic/market.html", {
        "active_page": "market", "competitors": competitors, "form": form,
    })


@login_required
def competitor_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_competitor"):
        _obj = get_object_or_404(Competitor, pk=pk)
        _label = str(_obj)
        _obj.delete()
        _log_action(request, "DELETE Competitor", _label)
    return redirect("strategic:market")


# ---------------- PESTEL ----------------

def pestel(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_pestelfactor" if obj_id else "strategic.add_pestelfactor"
        if _has_perm(request, perm):
            instance = get_object_or_404(PestelFactor, pk=obj_id) if obj_id else None
            form = PestelFactorForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE PestelFactor" if obj_id else "CREATE PestelFactor", str(form.instance))
                return redirect("strategic:pestel")
        else:
            form = PestelFactorForm()
    else:
        form = PestelFactorForm()

    LETTERS = {"political": "P", "economic": "E", "social": "S",
               "technological": "T", "environmental": "E", "legal": "L"}

    factors = PestelFactor.objects.all().prefetch_related("swot_items__business_unit", "cross_impact_factors", "linked_legal_requirements", "related_stakeholders")
    swot_code_map = _swot_code_map()
    for f in factors:
        for si in f.swot_items.all():
            si.swot_code = swot_code_map.get(si.pk, "")
    grouped = []
    for key, label in PestelFactor.CATEGORY_CHOICES:
        color, soft, icon = PestelFactor.CATEGORY_STYLE[key]
        cat_items = [f for f in factors if f.category == key]
        grouped.append({
            "key": key, "label": label, "color": color, "soft": soft, "icon": icon,
            "letter": LETTERS[key],
            "items": cat_items,
            "summary": cat_items,
        })

    top_factors = sorted(factors, key=lambda f: f.priority_score, reverse=True)[:8]

    return render(request, "strategic/pestel.html", {
        "active_page": "pestel", "grouped": grouped, "form": form, "top_factors": top_factors,
    })


@login_required
def pestel_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_pestelfactor"):
        _obj = get_object_or_404(PestelFactor, pk=pk)
        _label = str(_obj)
        _obj.delete()
        _log_action(request, "DELETE PestelFactor", _label)
    return redirect("strategic:pestel")


# ---------------- تحلیل اثرات متقابل (MICMAC) ----------------

def cross_impact(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_crossimpactfactor" if obj_id else "strategic.add_crossimpactfactor"
        if _has_perm(request, perm):
            instance = get_object_or_404(CrossImpactFactor, pk=obj_id) if obj_id else None
            form = CrossImpactFactorForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE CrossImpactFactor" if obj_id else "CREATE CrossImpactFactor", str(form.instance))
                return redirect("strategic:cross_impact")
        else:
            form = CrossImpactFactorForm()
    else:
        form = CrossImpactFactorForm()

    factors = list(CrossImpactFactor.objects.all())
    quadrants = {key: [] for key, _ in CrossImpactFactor.QUADRANT_CHOICES}
    for f in factors:
        quadrants[f.quadrant].append(f)

    return render(request, "strategic/cross_impact.html", {
        "active_page": "cross_impact", "quadrants": quadrants, "form": form,
        "pestel_factors": PestelFactor.objects.all(),
    })


@login_required
def cross_impact_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_crossimpactfactor"):
        obj = get_object_or_404(CrossImpactFactor, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE CrossImpactFactor", label)
    return redirect("strategic:cross_impact")


def _compute_influence_dependence(factors):
    """برای هر عامل، اثرگذاری (مجموع سطر) و وابستگی (مجموع ستون) را از روی ماتریس مستقیم حساب می‌کند."""
    links = {(l.from_factor_id, l.to_factor_id): l.score for l in CrossImpactLink.objects.all()}
    results = []
    for f in factors:
        influence = sum(links.get((f.pk, other.pk), 0) for other in factors if other.pk != f.pk)
        dependence = sum(links.get((other.pk, f.pk), 0) for other in factors if other.pk != f.pk)
        results.append({"factor": f, "influence": influence, "dependence": dependence})
    return results


def _suggest_quadrant(influence, dependence, median_influence, median_dependence):
    if influence >= median_influence and dependence < median_dependence:
        return "driver"
    if influence >= median_influence and dependence >= median_dependence:
        return "relay"
    if influence < median_influence and dependence < median_dependence:
        return "watch"
    return "resultant"


@login_required
def cross_impact_matrix(request):
    factors = list(CrossImpactFactor.objects.all().order_by("order"))

    if request.method == "POST" and _has_perm(request, "strategic.change_crossimpactfactor"):
        action = request.POST.get("action")
        if action == "save_matrix":
            for f in factors:
                for g in factors:
                    if f.pk == g.pk:
                        continue
                    key = f"score_{f.pk}_{g.pk}"
                    val = request.POST.get(key)
                    if val is not None and val != "":
                        CrossImpactLink.objects.update_or_create(
                            from_factor=f, to_factor=g, defaults={"score": int(val)},
                        )
            _log_action(request, "UPDATE CrossImpactLink matrix", "ماتریس اثرات مستقیم به‌روزرسانی شد")
            return redirect("strategic:cross_impact_matrix")

        elif action == "apply_suggestions":
            results = _compute_influence_dependence(factors)
            influences = sorted(r["influence"] for r in results)
            dependences = sorted(r["dependence"] for r in results)
            n = len(results)
            median_influence = influences[n // 2] if n else 0
            median_dependence = dependences[n // 2] if n else 0
            for r in results:
                suggested = _suggest_quadrant(r["influence"], r["dependence"], median_influence, median_dependence)
                r["factor"].quadrant = suggested
                r["factor"].save(update_fields=["quadrant"])
            _log_action(request, "APPLY CrossImpact suggestions", "پیشنهادهای محاسبه‌شده اعمال شد")
            return redirect("strategic:cross_impact")

    results = _compute_influence_dependence(factors)
    influences = sorted(r["influence"] for r in results)
    dependences = sorted(r["dependence"] for r in results)
    n = len(results)
    median_influence = influences[n // 2] if n else 0
    median_dependence = dependences[n // 2] if n else 0
    for r in results:
        r["suggested"] = _suggest_quadrant(r["influence"], r["dependence"], median_influence, median_dependence)
        r["suggested_label"] = dict(CrossImpactFactor.QUADRANT_CHOICES)[r["suggested"]]
        r["current_label"] = dict(CrossImpactFactor.QUADRANT_CHOICES)[r["factor"].quadrant]
        r["changed"] = r["suggested"] != r["factor"].quadrant

    links = {(l.from_factor_id, l.to_factor_id): l.score for l in CrossImpactLink.objects.all()}
    matrix_rows = []
    for f in factors:
        cells = []
        for g in factors:
            if f.pk == g.pk:
                cells.append({"to_factor": g, "is_diag": True, "score": None})
            else:
                cells.append({"to_factor": g, "is_diag": False, "score": links.get((f.pk, g.pk), 0)})
        matrix_rows.append({"from_factor": f, "cells": cells})

    max_influence = max((r["influence"] for r in results), default=1) or 1
    max_dependence = max((r["dependence"] for r in results), default=1) or 1

    plot_left, plot_right = 74, 650
    plot_top, plot_bottom = 26, 396
    for r in results:
        r["plot_x"] = round(plot_left + (r["dependence"] / max_dependence) * (plot_right - plot_left), 1)
        r["plot_y"] = round(plot_bottom - (r["influence"] / max_influence) * (plot_bottom - plot_top), 1)
    plot_median_x = round(plot_left + (median_dependence / max_dependence) * (plot_right - plot_left), 1)
    plot_median_y = round(plot_bottom - (median_influence / max_influence) * (plot_bottom - plot_top), 1)

    return render(request, "strategic/cross_impact_matrix.html", {
        "active_page": "cross_impact", "factors": factors, "results": results, "matrix_rows": matrix_rows,
        "median_influence": median_influence, "median_dependence": median_dependence,
        "max_influence": max_influence, "max_dependence": max_dependence,
        "plot_median_x": plot_median_x, "plot_median_y": plot_median_y,
        "plot_left": plot_left, "plot_right": plot_right, "plot_top": plot_top, "plot_bottom": plot_bottom,
    })


# ---------------- Porter's Five Forces ----------------

def porter(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_porterforce" if obj_id else "strategic.add_porterforce"
        if _has_perm(request, perm):
            instance = get_object_or_404(PorterForce, pk=obj_id) if obj_id else None
            form = PorterForceForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE PorterForce" if obj_id else "CREATE PorterForce", str(form.instance))
                return redirect("strategic:porter")
        else:
            form = PorterForceForm()
    else:
        form = PorterForceForm()

    forces = PorterForce.objects.all().prefetch_related("swot_items__business_unit", "cross_impact_factors", "linked_stakeholders")
    swot_code_map = _swot_code_map()
    for f in forces:
        for si in f.swot_items.all():
            si.swot_code = swot_code_map.get(si.pk, "")
    grouped = []
    for key, label in PorterForce.FORCE_CHOICES:
        color, soft, icon = PorterForce.FORCE_STYLE[key]
        cat_items = [f for f in forces if f.force == key]
        grouped.append({
            "key": key, "label": label, "color": color, "soft": soft, "icon": icon,
            "items": cat_items,
        })

    segments, connectors = _wheel_geometry(len(grouped), r_out=194, r_in=82, content_r=142)
    wheel_wedges = [
        dict(group=g, path=seg["path"], content_x=seg["content_x"], content_y=seg["content_y"])
        for g, seg in zip(grouped, segments)
    ]

    return render(request, "strategic/porter.html", {
        "active_page": "porter", "grouped": grouped,
        "wheel_wedges": wheel_wedges, "wheel_connectors": connectors,
        "form": form,
    })


@login_required
def porter_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_porterforce"):
        _obj = get_object_or_404(PorterForce, pk=pk)
        _label = str(_obj)
        _obj.delete()
        _log_action(request, "DELETE PorterForce", _label)
    return redirect("strategic:porter")


# ---------------- McKinsey 7S ----------------

def mckinsey7s(request):
    for key, _ in McKinsey7S.COMPONENT_CHOICES:
        McKinsey7S.objects.get_or_create(component=key)

    if request.method == "POST":
        comp_id = request.POST.get("obj_id")
        if _has_perm(request, "strategic.change_mckinsey7s"):
            instance = get_object_or_404(McKinsey7S, pk=comp_id)
            form = McKinsey7SForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE McKinsey7S", str(instance))
                return redirect("strategic:mckinsey7s")

    components = list(McKinsey7S.objects.all().prefetch_related("swot_items__business_unit"))
    order = list(dict(McKinsey7S.COMPONENT_CHOICES).keys())
    components.sort(key=lambda c: order.index(c.component))
    for c in components:
        c.color, c.soft, c.icon, c.english, c.group = McKinsey7S.STYLE[c.component]
    center = next((c for c in components if c.group == "center"), None)
    hard_items = [c for c in components if c.group == "hard"]
    soft_items = [c for c in components if c.group == "soft"]

    # کد S1/W1/O1/T1 هیچ‌جا توی دیتابیس ذخیره نمی‌شه — از تابع مشترک می‌گیریمش تا با
    # صفحه‌ی SWOT و بقیه‌ی جاهایی که ازش استفاده می‌کنن یکی باشه.
    swot_code_map = _swot_code_map()
    for c in components:
        for si in c.swot_items.all():
            si.swot_code = swot_code_map.get(si.pk, "")

    return render(request, "strategic/mckinsey7s.html", {
        "active_page": "mckinsey7s", "components": components, "form": McKinsey7SForm(),
        "center": center, "hard_items": hard_items, "soft_items": soft_items,
    })


# ---------------- زنجیره ارزش پورتر (تحلیل محیطی) ----------------

def value_chain(request):
    for key, _ in ValueChainActivity.ACTIVITY_CHOICES:
        ValueChainActivity.objects.get_or_create(activity=key)

    if request.method == "POST":
        act_id = request.POST.get("obj_id")
        if _has_perm(request, "strategic.change_valuechainactivity"):
            instance = get_object_or_404(ValueChainActivity, pk=act_id)
            form = ValueChainActivityForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE ValueChainActivity", str(instance))
                return redirect("strategic:value_chain")

    activities = list(ValueChainActivity.objects.all().prefetch_related("swot_items__business_unit"))
    swot_code_map = _swot_code_map()
    for a in activities:
        for si in a.swot_items.all():
            si.swot_code = swot_code_map.get(si.pk, "")
    order = list(dict(ValueChainActivity.ACTIVITY_CHOICES).keys())
    activities.sort(key=lambda a: order.index(a.activity))
    primary = [a for a in activities if a.activity_type == "primary"]
    support = [a for a in activities if a.activity_type == "support"]

    return render(request, "strategic/value_chain.html", {
        "active_page": "value_chain", "primary": primary, "support": support, "form": ValueChainActivityForm(),
    })


# ---------------- Strategic map (BSC) ----------------

THEME_PALETTE = ["#0f8a6a", "#1183c9", "#7b5cd6", "#d08a1f", "#17a3a3", "#d6402f", "#8a5a44", "#5a6474"]


def stratmap(request):
    business_units = list(BusinessUnit.objects.all())
    bu_id = request.POST.get("business_unit") or request.GET.get("bu")
    current_bu = None
    if bu_id:
        current_bu = next((b for b in business_units if str(b.pk) == str(bu_id)), None)
    if not current_bu and business_units:
        current_bu = business_units[0]

    themes = list(StrategyTheme.objects.filter(business_unit=current_bu)) if current_bu else []
    theme_color = {}
    for i, t in enumerate(themes):
        t.color = THEME_PALETTE[i % len(THEME_PALETTE)]
        theme_color[t.pk] = t.color

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_strategicobjective" if obj_id else "strategic.add_strategicobjective"
        if _has_perm(request, perm):
            instance = get_object_or_404(StrategicObjective, pk=obj_id) if obj_id else None
            form = StrategicObjectiveForm(request.POST, instance=instance, business_unit=current_bu)
            if form.is_valid():
                obj = form.save(commit=False)
                if not obj_id and current_bu:
                    obj.business_unit = current_bu
                obj.save()
                form.save_m2m()
                _log_action(request, "UPDATE StrategicObjective" if obj_id else "CREATE StrategicObjective", str(obj))
                bu_param = f"?bu={current_bu.pk}" if current_bu else ""
                return redirect(reverse("strategic:stratmap") + bu_param)
        else:
            form = StrategicObjectiveForm(business_unit=current_bu)
    else:
        form = StrategicObjectiveForm(business_unit=current_bu)

    objectives = list(
        StrategicObjective.objects.filter(business_unit=current_bu).select_related("theme", "source_tows").prefetch_related("feeds_into", "linked_kpis", "linked_operational_kpis", "source_tows__source_items")
        if current_bu else StrategicObjective.objects.none()
    )
    for o in objectives:
        o.theme_color = theme_color.get(o.theme_id, "var(--ink-faint)")
        o.theme_name = o.theme.name if o.theme_id else "بدون محور"

    PERSP_KEYS = {"financial": "fin", "customer": "cust", "process": "proc", "learning": "learn"}
    bands = []
    for p_key, p_label in StrategicObjective.PERSPECTIVE_CHOICES:
        bands.append({
            "key": p_key, "css": PERSP_KEYS[p_key], "label": p_label,
            "nodes": [o for o in objectives if o.perspective == p_key],
        })

    links = []
    for o in objectives:
        for target in o.feeds_into.all():
            links.append([f"obj-{o.pk}", f"obj-{target.pk}"])

    org_identity, _ = OrgIdentity.objects.get_or_create(pk=1)

    def _pct_color(pct):
        if pct is None:
            return "#9aa1ab"
        if pct < 80:
            return "#B0413E"
        if pct < 90:
            return "#C97A2B"
        return "#3E7A52"

    kpis_by_objective = {}
    circles_by_objective = {}
    for k in StrategicKPI.objects.filter(objective__in=objectives).select_related("objective"):
        kpis_by_objective.setdefault(k.objective_id, []).append({
            "type": "custom", "id": k.pk, "name": k.name, "unit": k.unit, "target": k.target, "actual": k.actual,
            "trend": k.trend, "status": k.status, "owner": k.owner, "period": k.period,
        })
        circles_by_objective.setdefault(k.objective_id, []).append({
            "name": k.name, "target": k.target, "actual": k.actual, "unit": k.unit,
            "pct": k.progress_pct, "color": _pct_color(k.progress_pct),
        })

    for o in objectives:
        for k in o.linked_kpis.all():
            kpis_by_objective.setdefault(o.pk, []).append({
                "type": "shared", "id": k.pk, "name": f"{k.code} — {k.name}", "unit": k.unit,
                "target": k.target_1405, "actual": k.actual_1405,
                "trend": "flat", "status": "on_track", "owner": "", "period": "",
            })
            circles_by_objective.setdefault(o.pk, []).append({
                "name": f"{k.code} — {k.name}", "target": k.target_1405, "actual": k.actual_1405, "unit": k.unit,
                "pct": k.manual_progress_value, "color": _pct_color(k.manual_progress_value),
            })
        for k in o.linked_operational_kpis.all():
            kpis_by_objective.setdefault(o.pk, []).append({
                "type": "operational", "id": k.pk, "name": f"{k.code} — {k.title}", "unit": k.unit,
                "target": k.target_1405, "actual": k.actual_1405,
                "trend": "flat", "status": "on_track", "owner": k.department, "period": "",
            })
            circles_by_objective.setdefault(o.pk, []).append({
                "name": f"{k.code} — {k.title}", "target": k.target_1405, "actual": k.actual_1405, "unit": k.unit,
                "pct": k.manual_progress_value, "color": _pct_color(k.manual_progress_value),
            })

    for o in objectives:
        o.kpi_count = len(kpis_by_objective.get(o.pk, []))
        o.kpi_circles = circles_by_objective.get(o.pk, [])

    return render(request, "strategic/stratmap.html", {
        "active_page": "stratmap", "bands": bands, "form": form,
        "links": links, "business_units": business_units, "current_bu": current_bu,
        "themes": themes, "theme_form": StrategyThemeForm(), "org_vision": org_identity.vision,
        "kpis_by_objective": kpis_by_objective, "kpi_form": StrategicKPIForm(),
    })


@login_required
def strategic_kpi_save(request):
    if request.method == "POST":
        obj_id = request.POST.get("kpi_id")
        objective_id = request.POST.get("objective_id")
        perm = "strategic.change_strategickpi" if obj_id else "strategic.add_strategickpi"
        if _has_perm(request, perm):
            instance = get_object_or_404(StrategicKPI, pk=obj_id) if obj_id else None
            form = StrategicKPIForm(request.POST, instance=instance)
            if form.is_valid():
                kpi = form.save(commit=False)
                if not obj_id:
                    kpi.objective_id = objective_id
                kpi.save()
                _log_action(request, "UPDATE StrategicKPI" if obj_id else "CREATE StrategicKPI", str(kpi))
    ref = request.POST.get("next") or reverse("strategic:stratmap")
    return redirect(ref)


@login_required
def strategic_kpi_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_strategickpi"):
        obj = get_object_or_404(StrategicKPI, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE StrategicKPI", label)
    ref = request.POST.get("next") or reverse("strategic:stratmap")
    return redirect(ref)


def stratmap_print(request):
    business_units = list(BusinessUnit.objects.all())
    bu_id = request.GET.get("bu")
    current_bu = None
    if bu_id:
        current_bu = next((b for b in business_units if str(b.pk) == str(bu_id)), None)
    if not current_bu and business_units:
        current_bu = business_units[0]

    objectives = list(
        StrategicObjective.objects.filter(business_unit=current_bu).select_related("theme").prefetch_related("feeds_into")
        if current_bu else StrategicObjective.objects.none()
    )
    links = []
    for o in objectives:
        o.theme_name = o.theme.name if o.theme_id else ""
        targets = list(o.feeds_into.all())
        o.feeds_codes = [t.code for t in targets]
        for t in targets:
            links.append([f"pcard-{o.pk}", f"pcard-{t.pk}"])

    bands = []
    PERSP_KEYS = {"financial": "fin", "customer": "cust", "process": "proc", "learning": "learn"}
    for p_key, p_label in StrategicObjective.PERSPECTIVE_CHOICES:
        bands.append({"label": p_label, "css": PERSP_KEYS[p_key], "nodes": [o for o in objectives if o.perspective == p_key]})

    return render(request, "strategic/stratmap_print.html", {"current_bu": current_bu, "bands": bands, "links": links})


@login_required
def stratmap_print_full(request):
    """خروجی کامل تک‌صفحه‌ای نقشه استراتژیک — اندازه‌ی صفحه دینامیک، بدون تقسیم به چند برگه."""
    business_units = list(BusinessUnit.objects.all())
    bu_id = request.GET.get("bu")
    current_bu = None
    if bu_id:
        current_bu = next((b for b in business_units if str(b.pk) == str(bu_id)), None)
    if not current_bu and business_units:
        current_bu = business_units[0]

    objectives = list(
        StrategicObjective.objects.filter(business_unit=current_bu).select_related("theme").prefetch_related("feeds_into")
        if current_bu else StrategicObjective.objects.none()
    )
    links = []
    for o in objectives:
        o.theme_name = o.theme.name if o.theme_id else ""
        targets = list(o.feeds_into.all())
        o.feeds_codes = [t.code for t in targets]
        for t in targets:
            links.append([f"pcard-{o.pk}", f"pcard-{t.pk}"])

    bands = []
    PERSP_KEYS = {"financial": "fin", "customer": "cust", "process": "proc", "learning": "learn"}
    for p_key, p_label in StrategicObjective.PERSPECTIVE_CHOICES:
        bands.append({"label": p_label, "css": PERSP_KEYS[p_key], "nodes": [o for o in objectives if o.perspective == p_key]})

    return render(request, "strategic/stratmap_print_full.html", {"current_bu": current_bu, "bands": bands, "links": links})


@login_required
def theme_add(request):
    if request.method == "POST" and _has_perm(request, "strategic.add_strategytheme"):
        bu_id = request.POST.get("business_unit")
        bu = get_object_or_404(BusinessUnit, pk=bu_id)
        name = request.POST.get("name", "").strip()
        if name:
            StrategyTheme.objects.create(business_unit=bu, name=name, order=bu.themes.count())
            _log_action(request, "CREATE StrategyTheme", f"{bu.name} — {name}")
        return redirect(reverse("strategic:stratmap") + f"?bu={bu.pk}")
    return redirect("strategic:stratmap")


@login_required
def theme_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_strategytheme"):
        _obj = get_object_or_404(StrategyTheme, pk=pk)
        bu_pk = _obj.business_unit_id
        _label = str(_obj)
        _obj.delete()
        _log_action(request, "DELETE StrategyTheme", _label)
        return redirect(reverse("strategic:stratmap") + f"?bu={bu_pk}")
    return redirect("strategic:stratmap")


@login_required
def objective_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_strategicobjective"):
        _obj = get_object_or_404(StrategicObjective, pk=pk)
        _label = str(_obj)
        bu_pk = _obj.business_unit_id
        _obj.delete()
        _log_action(request, "DELETE StrategicObjective", _label)
        if bu_pk:
            return redirect(reverse("strategic:stratmap") + f"?bu={bu_pk}")
    return redirect("strategic:stratmap")


@login_required
def business_unit_add(request):
    if request.method == "POST" and _has_perm(request, "strategic.add_businessunit"):
        name = request.POST.get("name", "").strip()
        archetype = request.POST.get("archetype", "other")
        next_page = request.POST.get("next", "strategic:stratmap")
        if name:
            bu = BusinessUnit.objects.create(name=name, archetype=archetype, order=BusinessUnit.objects.count())
            _log_action(request, "CREATE BusinessUnit", name)
            return redirect(reverse(next_page) + f"?bu={bu.pk}")
        return redirect(next_page)
    return redirect("strategic:stratmap")


# ---------------- SWOT ----------------

@login_required
def swot(request):
    business_units = list(BusinessUnit.objects.all())
    bu_id = request.POST.get("business_unit") or request.GET.get("bu")
    current_bu = None
    if bu_id:
        current_bu = next((b for b in business_units if str(b.pk) == str(bu_id)), None)
    if not current_bu and business_units:
        current_bu = business_units[0]
    bu_param = f"?bu={current_bu.pk}" if current_bu else ""

    if request.method == "POST":
        cat = request.POST.get("category", "")
        obj_id = request.POST.get("obj_id")
        if cat in ("s", "w", "o", "t"):
            perm = "strategic.change_swotitem" if obj_id else "strategic.add_swotitem"
            if _has_perm(request, perm):
                instance = get_object_or_404(SWOTItem, pk=obj_id) if obj_id else None
                form = SWOTItemForm(request.POST, instance=instance)
                if form.is_valid():
                    obj = form.save(commit=False)
                    if not obj_id:
                        obj.business_unit = current_bu
                    obj.save()
                    _log_action(request, "UPDATE SWOTItem" if obj_id else "CREATE SWOTItem", str(obj))
                    return redirect(reverse("strategic:swot") + bu_param)
        elif cat in ("so", "st", "wo", "wt"):
            perm = "strategic.change_towsstrategy" if obj_id else "strategic.add_towsstrategy"
            if _has_perm(request, perm):
                t_instance = get_object_or_404(TOWSStrategy, pk=obj_id) if obj_id else None
                tform = TOWSStrategyForm(request.POST, instance=t_instance, business_unit=current_bu)
                if tform.is_valid():
                    tobj = tform.save(commit=False)
                    if not obj_id:
                        tobj.business_unit = current_bu
                    tobj.save()
                    tform.save_m2m()
                    _log_action(request, "UPDATE TOWSStrategy" if obj_id else "CREATE TOWSStrategy", str(tobj))
                    return redirect(reverse("strategic:swot") + bu_param)

    if current_bu:
        s_items = list(SWOTItem.objects.filter(category="s", business_unit=current_bu))
        w_items = list(SWOTItem.objects.filter(category="w", business_unit=current_bu))
        o_items = list(SWOTItem.objects.filter(category="o", business_unit=current_bu))
        t_items = list(SWOTItem.objects.filter(category="t", business_unit=current_bu))
    else:
        s_items = w_items = o_items = t_items = []

    def avg_w(items):
        return round(sum(i.weight for i in items) / len(items), 1) if items else 0

    s_score, w_score, o_score, t_score = avg_w(s_items), avg_w(w_items), avg_w(o_items), avg_w(t_items)

    def norm(score):
        return (score - 3) / 2 if score else 0

    internal = norm(s_score) - norm(w_score)
    external = norm(o_score) - norm(t_score)
    pos_x = 50 + internal * 42
    pos_y = 50 - external * 42
    if internal >= 0 and external >= 0:
        posture, posture_color = "راهبرد تهاجمی (SO)", "var(--s)"
        posture_desc = "کسب‌وکار در موقعیت قدرتمندی قرار دارد؛ باید با تکیه بر قوت‌ها، حداکثر بهره را از فرصت‌ها ببرد و بر رشد و توسعه تمرکز کند."
    elif internal >= 0 and external < 0:
        posture, posture_color = "راهبرد تنوع (ST)", "var(--w)"
        posture_desc = "کسب‌وکار با وجود تهدیدهای بیرونی، از قوت داخلی کافی برخوردار است؛ باید این قوت‌ها را برای خنثی‌سازی تهدیدها یا تنوع‌بخشی به‌کار گیرد."
    elif internal < 0 and external >= 0:
        posture, posture_color = "راهبرد بازنگری (WO)", "var(--o)"
        posture_desc = "فرصت‌های بیرونی مناسبی وجود دارد، اما ضعف‌های داخلی مانع بهره‌برداری کامل است؛ اولویت، رفع این ضعف‌هاست."
    else:
        posture, posture_color = "راهبرد تدافعی (WT)", "var(--t)"
        posture_desc = "کسب‌وکار هم‌زمان با ضعف داخلی و تهدید بیرونی روبروست؛ اولویت، کاهش آسیب‌پذیری و پرهیز از تصمیمات پرریسک است."

    tows = {}
    for key, _ in TOWSStrategy.CATEGORY_CHOICES:
        tows[key] = list(
            TOWSStrategy.objects.filter(category=key, business_unit=current_bu)
            .prefetch_related("source_items", "objectives")
        ) if current_bu else []

    swot_items_data = []
    for letter, items in (("S", s_items), ("W", w_items), ("O", o_items), ("T", t_items)):
        for i, it in enumerate(items, start=1):
            it.code = f"{letter}{i}"
            swot_items_data.append({"id": it.pk, "cat": it.category, "code": it.code, "text": it.text})

    return render(request, "strategic/swot.html", {
        "active_page": "swot",
        "business_units": business_units, "current_bu": current_bu,
        "s_items": s_items, "w_items": w_items, "o_items": o_items, "t_items": t_items,
        "s_score": s_score, "w_score": w_score, "o_score": o_score, "t_score": t_score,
        "pos_x": pos_x, "pos_y": pos_y, "posture": posture, "posture_color": posture_color, "posture_desc": posture_desc,
        "internal_dominant": "قوت‌ها بر ضعف‌ها" if internal >= 0 else "ضعف‌ها بر قوت‌ها",
        "external_dominant": "فرصت‌ها بر تهدیدها" if external >= 0 else "تهدیدها بر فرصت‌ها",
        "tows": tows, "swot_items_data": swot_items_data,
        "form": SWOTItemForm(),
        "tows_form": TOWSStrategyForm(business_unit=current_bu),
        "pestel_factors": PestelFactor.objects.all(),
        "porter_forces": PorterForce.objects.all(),
        "stakeholders_list": Stakeholder.objects.all(),
        "scenarios_list": Scenario.objects.all(),
        "s7_components": McKinsey7S.objects.all(),
        "vc_activities": ValueChainActivity.objects.all(),
    })


@login_required
def swot_print(request):
    business_units = list(BusinessUnit.objects.all())
    bu_id = request.GET.get("bu")
    current_bu = None
    if bu_id:
        current_bu = next((b for b in business_units if str(b.pk) == str(bu_id)), None)
    if not current_bu and business_units:
        current_bu = business_units[0]

    if current_bu:
        s_items = list(SWOTItem.objects.filter(category="s", business_unit=current_bu))
        w_items = list(SWOTItem.objects.filter(category="w", business_unit=current_bu))
        o_items = list(SWOTItem.objects.filter(category="o", business_unit=current_bu))
        t_items = list(SWOTItem.objects.filter(category="t", business_unit=current_bu))
        tows = {key: list(TOWSStrategy.objects.filter(category=key, business_unit=current_bu))
                for key, _ in TOWSStrategy.CATEGORY_CHOICES}
    else:
        s_items = w_items = o_items = t_items = []
        tows = {}

    return render(request, "strategic/swot_print.html", {
        "current_bu": current_bu,
        "s_items": s_items, "w_items": w_items, "o_items": o_items, "t_items": t_items,
        "tows": tows,
    })


@login_required
def swot_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_swotitem"):
        _obj = get_object_or_404(SWOTItem, pk=pk)
        _label = str(_obj)
        bu_pk = _obj.business_unit_id
        _obj.delete()
        _log_action(request, "DELETE SWOTItem", _label)
        if bu_pk:
            return redirect(reverse("strategic:swot") + f"?bu={bu_pk}")
    return redirect("strategic:swot")


@login_required
def tows_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_towsstrategy"):
        _obj = get_object_or_404(TOWSStrategy, pk=pk)
        _label = str(_obj)
        bu_pk = _obj.business_unit_id
        _obj.delete()
        _log_action(request, "DELETE TOWSStrategy", _label)
        if bu_pk:
            return redirect(reverse("strategic:swot") + f"?bu={bu_pk}")
    return redirect("strategic:swot")


# ---------------- Risk register ----------------

@login_required
def risk(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_risk" if obj_id else "strategic.add_risk"
        if _has_perm(request, perm):
            instance = get_object_or_404(Risk, pk=obj_id) if obj_id else None
            form = RiskForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE Risk" if obj_id else "CREATE Risk", str(form.instance))
                return redirect("strategic:risk")
        else:
            form = RiskForm()
    else:
        form = RiskForm()

    risks = list(Risk.objects.all().prefetch_related("related_scenario"))
    risks_sorted = sorted(risks, key=lambda r: -r.residual_score)
    for idx, r in enumerate(risks_sorted, start=1):
        r.display_no = idx  # runtime-only, used for bubble/row numbering

    ZONE_COLOR = {"low": "#2fa96b", "med": "#e3b23c", "high": "#e2792e", "crit": "#d6402f"}

    # ماتریس ۵×۵ (احتمال × شدت اثر) — چون احتمال/اثر می‌توانند اعشاری باشند (طبق فایل رسمی)،
    # هر ریسک برای «جایگذاری روی خانه‌ی ماتریس» به نزدیک‌ترین عدد صحیح ۱ تا ۵ گرد می‌شود؛
    # امتیاز واقعی (RPN) و مقدار دقیق اعشاری در همه‌جای دیگر (جدول، هاور) دست‌نخورده می‌مانند.
    def _bucket(value):
        return min(5, max(1, round(float(value))))

    matrix = []
    for impact in range(5, 0, -1):
        row = []
        for likelihood in range(1, 6):
            s = likelihood * impact
            zone = Risk._zone_of(s)
            cell_risks = [r for r in risks_sorted if _bucket(r.likelihood) == likelihood and _bucket(r.impact) == impact]
            row.append({"zone": zone, "color": ZONE_COLOR[zone], "score": s,
                        "likelihood": likelihood, "impact": impact, "risks": cell_risks})
        matrix.append(row)

    top_risks = risks_sorted[:5]

    categories = []
    for key, label in Risk.CATEGORY_CHOICES:
        n = sum(1 for r in risks if r.category == key)
        categories.append({"key": key, "label": label, "color": Risk.CATEGORY_COLOR[key], "count": n})
    max_cat = max([c["count"] for c in categories], default=0) or 1
    for c in categories:
        c["pct"] = round(c["count"] / max_cat * 100)

    total = len(risks)
    high_or_crit = sum(1 for r in risks if r.residual_score >= 10)
    above_appetite = sum(1 for r in risks if r.residual_score > 9)
    avg_score = round(sum(r.residual_score for r in risks) / total, 1) if total else 0

    return render(request, "strategic/risk.html", {
        "active_page": "risk", "matrix": matrix, "risks": risks_sorted, "top_risks": top_risks,
        "form": form, "categories": categories, "max_cat": max_cat,
        "total": total, "high_or_crit": high_or_crit, "above_appetite": above_appetite,
        "avg_score": avg_score,
    })


@login_required
def risk_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_risk"):
        _obj = get_object_or_404(Risk, pk=pk)
        _label = str(_obj)
        _obj.delete()
        _log_action(request, "DELETE Risk", _label)
    return redirect("strategic:risk")


# ---------------- چشم‌انداز، مأموریت و ارزش‌های سازمانی ----------------
# ویرایش این بخش فقط برای ادمین اصلی (Superuser) مجاز است، نه گروه «ویرایشگران».

def _wheel_geometry(n, cx=200, cy=200, r_out=194, r_in=82, gap_deg=1.6, content_r=142):
    """هندسه‌ی چرخ SVG (پای‌چارت واقعی، نه چیدمان دایره‌ای CSS): برای n قطعه‌ی بیرونی،
    مسیر هر قطعه (donut sector) + مختصات محتوای هر قطعه + نقاط اتصال بین قطعات را
    برمی‌گرداند. زاویه‌ی صفر = بالای چرخ (۱۲)، در جهت عقربه‌های ساعت."""
    if n <= 0:
        return [], []

    def polar(r, deg):
        rad = math.radians(deg)
        return cx + r * math.sin(rad), cy - r * math.cos(rad)

    step = 360.0 / n
    segments = []
    for i in range(n):
        start = i * step + gap_deg / 2
        end = (i + 1) * step - gap_deg / 2
        x1, y1 = polar(r_out, start)
        x2, y2 = polar(r_out, end)
        x3, y3 = polar(r_in, end)
        x4, y4 = polar(r_in, start)
        large_arc = 1 if (end - start) > 180 else 0
        path = (
            f"M {x1:.2f} {y1:.2f} A {r_out} {r_out} 0 {large_arc} 1 {x2:.2f} {y2:.2f} "
            f"L {x3:.2f} {y3:.2f} A {r_in} {r_in} 0 {large_arc} 0 {x4:.2f} {y4:.2f} Z"
        )
        mid = (start + end) / 2
        content_x, content_y = polar(content_r, mid)
        arrow_x, arrow_y = polar(r_in + 14, mid)
        arrow_deg = mid + 180
        segments.append(dict(
            path=path, content_x=round(content_x, 1), content_y=round(content_y, 1),
            arrow_x=round(arrow_x, 1), arrow_y=round(arrow_y, 1), arrow_deg=round(arrow_deg, 1),
        ))

    connectors = []
    for i in range(n):
        x, y = polar(r_out, i * step)
        connectors.append(dict(x=round(x, 1), y=round(y, 1)))

    return segments, connectors


# استانداردهای مدیریتی — نمایشی/ثابت، مطابق پوستر رسمی خط‌مشی سیستم‌های مدیریتی سایپا یدک.
ORG_ISO_STANDARDS = [
    dict(code="ISO 9001:2015", title="مدیریت کیفیت", domain="کلیه فرایندها"),
    dict(code="ISO 10002:2018", title="مدیریت شکایات مشتریان", domain="خدمات پس از فروش"),
    dict(code="ISO 10004:2018", title="اندازه‌گیری رضایت مشتری", domain="بازاریابی و فروش"),
    dict(code="ISO 10015:2019", title="مدیریت آموزش", domain="توسعه منابع انسانی"),
    dict(code="ISO 14001:2015", title="مدیریت زیست‌محیطی", domain="تمامی فعالیت‌ها"),
]

# زنجیره‌ی ارزش ← هدف استراتژیک ← شاخص کلیدی ← پروژه‌های مرتبط — صرفاً نمونه‌ی نمایشی برای
# توضیح نوع ارتباط؛ به داده‌ی واقعی نقشه استراتژیک/KPI وصل نیست (طبق تصمیم صریح کاربر).
ORG_VALUE_CHAIN_SAMPLES = [
    dict(value="مشتری‌مداری", goal="افزایش رضایت مشتریان از خدمات پس از فروش",
         kpi="شاخص رضایت مشتری (CSI)", project="پلتفرم یکپارچه مدیریت ارتباط با مشتری (CRM)"),
]


def org_identity(request):
    identity, _ = OrgIdentity.objects.get_or_create(pk=1)

    if request.method == "POST" and request.user.is_superuser:
        form_kind = request.POST.get("form_kind")

        if form_kind == "identity":
            identity.vision = request.POST.get("vision", "").strip()
            identity.mission = request.POST.get("mission", "").strip()
            identity.management_statement = request.POST.get("management_statement", "").strip()
            identity.signed_by = request.POST.get("signed_by", "").strip()
            identity.signed_role = request.POST.get("signed_role", "").strip()
            identity.signed_date = request.POST.get("signed_date", "").strip()
            identity.save()
            _log_action(request, "UPDATE OrgIdentity", "چشم‌انداز/مأموریت")
            return redirect("strategic:org_identity")

        elif form_kind == "value":
            obj_id = request.POST.get("obj_id")
            text = request.POST.get("text", "").strip()
            is_center = bool(request.POST.get("is_center"))
            icon = request.POST.get("icon", "").strip() or "fa-solid fa-star"
            color = request.POST.get("color", "").strip() or "#C97A2B"
            definition = request.POST.get("definition", "").strip()
            expected_behaviors = request.POST.get("expected_behaviors", "").strip()
            examples = request.POST.get("examples", "").strip()
            related_policy_id = request.POST.get("related_policy") or None
            related_objective_ids = request.POST.getlist("related_objectives")
            if text:
                field_values = dict(
                    text=text, is_center=is_center, icon=icon, color=color,
                    definition=definition, expected_behaviors=expected_behaviors, examples=examples,
                    related_policy_id=related_policy_id,
                )
                if obj_id:
                    OrgValue.objects.filter(pk=obj_id).update(**field_values)
                    value_obj = OrgValue.objects.get(pk=obj_id)
                else:
                    value_obj = OrgValue.objects.create(order=OrgValue.objects.count(), **field_values)
                value_obj.related_objectives.set(related_objective_ids)
                _log_action(request, "UPDATE OrgValue" if obj_id else "CREATE OrgValue", text)
            return redirect("strategic:org_identity")

        elif form_kind == "policy":
            obj_id = request.POST.get("obj_id")
            number = request.POST.get("number") or 0
            text = request.POST.get("text", "").strip()
            if text:
                if obj_id:
                    QualityPolicyPoint.objects.filter(pk=obj_id).update(number=number, text=text)
                else:
                    QualityPolicyPoint.objects.create(number=number, text=text, order=QualityPolicyPoint.objects.count())
                _log_action(request, "UPDATE QualityPolicyPoint" if obj_id else "CREATE QualityPolicyPoint", text)
            return redirect("strategic:org_identity")

    values = list(OrgValue.objects.all())
    outer_values = [v for v in values if not v.is_center]
    center_value = next((v for v in values if v.is_center), None)
    policy_points = list(QualityPolicyPoint.objects.all())

    segments, connectors = _wheel_geometry(len(outer_values))
    wheel_wedges = [
        dict(value=v, path=seg["path"], content_x=seg["content_x"], content_y=seg["content_y"])
        for v, seg in zip(outer_values, segments)
    ]

    def _obj_summary(o):
        kpis_qs = list(o.kpis.all())
        kpi_labels = [f"{k.code} — {k.name}" for k in kpis_qs]
        # روی هدف کلان، پروژه مستقیم وصل نیست — از طریق شاخص کلان می‌رسیم
        initiatives = []
        seen_init_pks = set()
        for k in kpis_qs:
            for i in k.initiatives.all():
                if i.pk not in seen_init_pks:
                    seen_init_pks.add(i.pk)
                    initiatives.append(i.title)
        return {
            "id": o.pk, "code": o.code, "title": o.title,
            "business_unit": "",
            "kpis": kpi_labels, "initiatives": initiatives,
        }

    values_for_js = {
        str(v.pk): {
            "text": v.text, "icon": v.icon, "color": v.color, "is_center": v.is_center,
            "policy_number": v.related_policy.number if v.related_policy else None,
            "policy_text": v.related_policy.text if v.related_policy else None,
            "objectives": [_obj_summary(o) for o in v.related_objectives.prefetch_related(
                "kpis", "kpis__initiatives"
            )],
        } for v in values
    }

    objective_choices = [
        {"id": o.pk, "code": o.code, "title": o.title, "bu": ""}
        for o in CompanyObjective.objects.all()
    ]

    return render(request, "strategic/org_identity.html", {
        "active_page": "org_identity", "identity": identity,
        "outer_values": outer_values, "center_value": center_value,
        "policy_points": policy_points,
        "wheel_wedges": wheel_wedges, "wheel_connectors": connectors,
        "values_for_js": values_for_js, "objective_choices": objective_choices,
        "iso_standards": ORG_ISO_STANDARDS,
        "value_chain_samples": ORG_VALUE_CHAIN_SAMPLES,
        "icon_choices": OrgValue.ICON_CHOICES,
        "color_choices": OrgValue.COLOR_CHOICES,
    })


def org_value_delete(request, pk):
    if request.method == "POST" and request.user.is_superuser:
        obj = get_object_or_404(OrgValue, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE OrgValue", label)
    return redirect("strategic:org_identity")


def policy_point_delete(request, pk):
    if request.method == "POST" and request.user.is_superuser:
        obj = get_object_or_404(QualityPolicyPoint, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE QualityPolicyPoint", label)
    return redirect("strategic:org_identity")


# ---------------- سناریوهای راهبردی ----------------

def scenarios(request):
    for key, _ in Scenario.QUADRANT_CHOICES:
        Scenario.objects.get_or_create(quadrant=key)
    axes, _ = ScenarioAxes.objects.get_or_create(pk=1)

    if request.method == "POST":
        form_kind = request.POST.get("form_kind")
        if form_kind == "scenario":
            obj_id = request.POST.get("obj_id")
            if _has_perm(request, "strategic.change_scenario"):
                instance = get_object_or_404(Scenario, pk=obj_id)
                form = ScenarioForm(request.POST, instance=instance)
                if form.is_valid():
                    if form.cleaned_data.get("is_selected"):
                        Scenario.objects.exclude(pk=instance.pk).update(is_selected=False)
                    form.save()
                    # برچسب‌های دستی متن (تکه‌متن ← عامل ماتریس) — از سمت فرانت به‌شکل JSON میان
                    try:
                        raw = json.loads(request.POST.get("highlights_json", "[]"))
                    except (ValueError, TypeError):
                        raw = []
                    instance.highlights.all().delete()
                    valid_factor_ids = set(CrossImpactFactor.objects.values_list("pk", flat=True))
                    new_highlights = []
                    for i, item in enumerate(raw):
                        phrase = str(item.get("phrase", "")).strip()
                        factor_id = item.get("factor_id")
                        if not phrase or factor_id not in valid_factor_ids:
                            continue
                        new_highlights.append(ScenarioHighlight(
                            scenario=instance, cross_impact_factor_id=factor_id, phrase=phrase, order=i,
                        ))
                    if new_highlights:
                        ScenarioHighlight.objects.bulk_create(new_highlights)
                    _log_action(request, "UPDATE Scenario", str(instance))
                    return redirect("strategic:scenarios")
        elif form_kind == "axes":
            if _has_perm(request, "strategic.change_scenarioaxes"):
                form = ScenarioAxesForm(request.POST, instance=axes)
                if form.is_valid():
                    form.save()
                    _log_action(request, "UPDATE ScenarioAxes", "محورهای سناریو")
                    return redirect("strategic:scenarios")
        elif form_kind == "response_strategy":
            scenario_id = request.POST.get("scenario_id")
            obj_id = request.POST.get("obj_id")
            perm = "strategic.change_scenarioresponsestrategy" if obj_id else "strategic.add_scenarioresponsestrategy"
            if _has_perm(request, perm):
                scenario_obj = get_object_or_404(Scenario, pk=scenario_id)
                instance = get_object_or_404(ScenarioResponseStrategy, pk=obj_id) if obj_id else None
                form = ScenarioResponseStrategyForm(request.POST, instance=instance)
                if form.is_valid():
                    strategy = form.save(commit=False)
                    strategy.scenario = scenario_obj
                    if not obj_id:
                        strategy.order = scenario_obj.response_strategies.count() + 1
                    strategy.save()
                    _log_action(request, "UPDATE ScenarioResponseStrategy" if obj_id else "CREATE ScenarioResponseStrategy", str(strategy))
                    return redirect(f"{reverse('strategic:scenarios')}?open_strategies={scenario_obj.quadrant}")
                else:
                    messages.error(request, f"ثبت راهبرد ناموفق بود: {form.errors.as_text()}")

    scenario_map = {s.quadrant: s for s in Scenario.objects.all().prefetch_related("swot_items__business_unit", "risks", "highlights__cross_impact_factor")}
    swot_code_map = _swot_code_map()
    for s in scenario_map.values():
        for si in s.swot_items.all():
            si.swot_code = swot_code_map.get(si.pk, "")
    business_units = list(BusinessUnit.objects.all())
    foggy_tooltip = "\n".join(f"{bu.name} (فرصت و تهدید)" for bu in business_units)

    cross_impact_factors = list(CrossImpactFactor.objects.all().order_by("quadrant", "order"))
    for s in scenario_map.values():
        s.highlights_json = json.dumps([
            {"phrase": h.phrase, "factor_id": h.cross_impact_factor_id} for h in s.highlights.all()
        ], ensure_ascii=False)

    return render(request, "strategic/scenarios.html", {
        "active_page": "scenarios", "scenario_map": scenario_map, "axes": axes,
        "scenario_form": ScenarioForm(), "axes_form": ScenarioAxesForm(instance=axes),
        "foggy_tooltip": foggy_tooltip, "cross_impact_factors": cross_impact_factors,
        "response_strategy_form": ScenarioResponseStrategyForm(),
    })


@login_required
def scenario_response_strategy_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_scenarioresponsestrategy"):
        obj = get_object_or_404(ScenarioResponseStrategy, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE ScenarioResponseStrategy", label)
    return redirect("strategic:scenarios")


# ---------------- اهداف کلان و شاخص‌های سطح کل شرکت ----------------

def company_goals(request):
    if request.method == "POST":
        form_kind = request.POST.get("form_kind")
        obj_id = request.POST.get("obj_id")

        if form_kind == "objective":
            perm = "strategic.change_companyobjective" if obj_id else "strategic.add_companyobjective"
            if _has_perm(request, perm):
                instance = get_object_or_404(CompanyObjective, pk=obj_id) if obj_id else None
                form = CompanyObjectiveForm(request.POST, instance=instance)
                if form.is_valid():
                    form.save()
                    _log_action(request, "UPDATE CompanyObjective" if obj_id else "CREATE CompanyObjective", str(form.instance))
                    return redirect("strategic:company_goals")

        elif form_kind == "kpi":
            perm = "strategic.change_companykpi" if obj_id else "strategic.add_companykpi"
            if _has_perm(request, perm):
                instance = get_object_or_404(CompanyKPI, pk=obj_id) if obj_id else None
                form = CompanyKPIForm(request.POST, instance=instance)
                if form.is_valid():
                    form.save()
                    _log_action(request, "UPDATE CompanyKPI" if obj_id else "CREATE CompanyKPI", str(form.instance))
                    return redirect("strategic:company_goals")

    kpis = list(CompanyKPI.objects.all().select_related("source_operational_kpi").prefetch_related("objectives"))
    objectives = list(CompanyObjective.objects.all().prefetch_related("kpis"))
    grouped = []
    seen_groups = {}
    for o in objectives:
        key = o.group_title
        if key not in seen_groups:
            seen_groups[key] = {"group_title": key, "objectives": []}
            grouped.append(seen_groups[key])
        seen_groups[key]["objectives"].append(o)

    operational_kpi_choices = [
        {
            "id": k.pk, "code": k.code, "title": k.title, "unit": k.unit, "domain": k.domain,
            "target_1404": k.target_1404, "actual_1404": k.actual_1404,
            "target_1405": k.target_1405, "actual_1405": k.actual_1405, "progress_1405": k.progress_1405,
        }
        for k in OperationalKPI.objects.all()
    ]

    return render(request, "strategic/company_goals.html", {
        "active_page": "company_goals", "grouped": grouped, "kpis": kpis,
        "objective_form": CompanyObjectiveForm(),
        "kpi_form": CompanyKPIForm(),
        "operational_kpi_choices": operational_kpi_choices,
    })


@login_required
def company_objective_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_companyobjective"):
        obj = get_object_or_404(CompanyObjective, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE CompanyObjective", label)
    return redirect("strategic:company_goals")


@login_required
@login_required
def company_kpi_refresh_from_source(request):
    if request.method != "POST" or not _has_perm(request, "strategic.change_companykpi"):
        messages.error(request, "این عملیات مجاز نیست.")
        return redirect("strategic:company_goals")

    linked_kpis = CompanyKPI.objects.filter(source_operational_kpi__isnull=False).select_related("source_operational_kpi")
    updated = 0
    for kpi in linked_kpis:
        src = kpi.source_operational_kpi
        # فقط ستون‌های عددی/واحد/حوزه به‌روزرسانی می‌شوند — نام (name) و اهداف مرتبط
        # (objectives) دست‌نخورده باقی می‌مانند، طبق تصمیم کاربر.
        kpi.domain = src.domain
        kpi.unit = src.unit
        kpi.target_1404 = src.target_1404
        kpi.actual_1404 = src.actual_1404
        kpi.target_1405 = src.target_1405
        kpi.actual_1405 = src.actual_1405
        kpi.progress_1405 = src.progress_1405
        kpi.save(update_fields=["domain", "unit", "target_1404", "actual_1404", "target_1405", "actual_1405", "progress_1405"])
        updated += 1

    _log_action(request, "REFRESH CompanyKPI from source", f"{updated} شاخص")
    if updated:
        messages.success(request, f"{updated} شاخص از بانک شاخص‌های عملیاتی به‌روزرسانی شد.")
    else:
        messages.warning(request, "هیچ شاخص کلانی به بانک عملیاتی وصل نیست — چیزی برای به‌روزرسانی نبود.")
    return redirect("strategic:company_goals")


def company_kpi_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_companykpi"):
        obj = get_object_or_404(CompanyKPI, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE CompanyKPI", label)
    return redirect("strategic:company_goals")


# ---------------- اسناد و دستورالعمل‌ها ----------------

@login_required
def documents(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_document" if obj_id else "strategic.add_document"
        if _has_perm(request, perm):
            instance = get_object_or_404(Document, pk=obj_id) if obj_id else None
            form = DocumentForm(request.POST, request.FILES, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE Document" if obj_id else "CREATE Document", str(form.instance))
                return redirect("strategic:documents")
        else:
            form = DocumentForm()
    else:
        form = DocumentForm()

    docs = Document.objects.all()
    upstream = [d for d in docs if d.category == "upstream"]
    guideline = [d for d in docs if d.category == "guideline"]

    return render(request, "strategic/documents.html", {
        "active_page": "documents", "upstream": upstream, "guideline": guideline, "form": form,
    })


@login_required
def document_download(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    try:
        f = doc.file.open("rb")
    except (FileNotFoundError, ValueError):
        raise Http404("فایل یافت نشد")
    is_pdf = doc.file_ext == "pdf"
    response = FileResponse(f, as_attachment=not is_pdf, filename=doc.file.name.split("/")[-1])
    return response


@login_required
def document_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_document"):
        obj = get_object_or_404(Document, pk=pk)
        label = str(obj)
        obj.file.delete(save=False)
        obj.delete()
        _log_action(request, "DELETE Document", label)
    return redirect("strategic:documents")


# ---------------- ورود/خروجی اکسل شاخص‌های کلیدی شرکت (فقط مدیر سیستم) ----------------

_KPI_EXCEL_HEADERS = [
    "کد", "حوزه (Q/D/C/M)", "شاخص", "واحد سنجش", "هدف سال گذشته", "عملکرد سال گذشته",
    "هدف سال جاری", "عملکرد سال جاری", "درصد تحقق", "صرفاً پایشی (بله/خیر)",
    "اهداف مرتبط (مثلاً O1-O2-O3)", "ملاحظات", "ترتیب نمایش",
]


@login_required
def company_kpi_export(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:company_goals")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "شاخص‌های کلیدی"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(_KPI_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_i, k in enumerate(CompanyKPI.objects.all().prefetch_related("objectives"), start=2):
        related = "-".join(o.code for o in k.objectives.all())
        values = [
            k.code, k.domain, k.name, k.unit, k.target_1404, k.actual_1404,
            k.target_1405, k.actual_1405, k.progress_1405, "بله" if k.is_monitoring else "خیر",
            related, k.notes, k.order,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_i, column=col, value=val)

    widths = [8, 14, 34, 12, 10, 12, 10, 12, 12, 14, 22, 24, 10]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT CompanyKPI Excel", f"{CompanyKPI.objects.count()} ردیف")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="shakhes-haye-kelidi.xlsx"'
    return response


@login_required
def company_kpi_import(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:company_goals")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "فایلی انتخاب نشده است.")
        return redirect("strategic:company_goals")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"], data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "فایل اکسل قابل خواندن نیست. لطفاً فرمت را بررسی کنید.")
        return redirect("strategic:company_goals")

    def _s(v):
        return "" if v is None else str(v).strip()

    created, updated, skipped = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code = _s(row[0])
        if not code:
            skipped += 1
            continue
        domain = _s(row[1])[:1].upper() or "Q"
        if domain not in dict(CompanyKPI.DOMAIN_CHOICES):
            domain = "Q"
        name = _s(row[2])
        unit = _s(row[3]) if len(row) > 3 else ""
        target_1404 = clean_number_string(_s(row[4]) if len(row) > 4 else "")
        actual_1404 = clean_number_string(_s(row[5]) if len(row) > 5 else "")
        target_1405 = clean_number_string(_s(row[6]) if len(row) > 6 else "")
        actual_1405 = clean_number_string(_s(row[7]) if len(row) > 7 else "")
        progress_1405 = clean_number_string(_s(row[8]) if len(row) > 8 else "")
        is_monitoring = _s(row[9]).startswith("بل") if len(row) > 9 else False
        related_raw = _s(row[10]) if len(row) > 10 else ""
        notes = _s(row[11]) if len(row) > 11 else ""
        try:
            order = int(row[12]) if len(row) > 12 and row[12] not in (None, "") else 0
        except (TypeError, ValueError):
            order = 0

        kpi, was_created = CompanyKPI.objects.update_or_create(
            code=code,
            defaults=dict(
                domain=domain, name=name, unit=unit,
                target_1404=target_1404, actual_1404=actual_1404,
                target_1405=target_1405, actual_1405=actual_1405,
                progress_1405=progress_1405, is_monitoring=is_monitoring,
                notes=notes, order=order,
            ),
        )
        related_codes = re.findall(r"O\d+", related_raw)
        if related_codes:
            objs = list(CompanyObjective.objects.filter(code__in=related_codes))
            kpi.objectives.set(objs)

        created += 1 if was_created else 0
        updated += 0 if was_created else 1

    _log_action(request, "IMPORT CompanyKPI Excel", f"{created} جدید، {updated} به‌روزشده، {skipped} رد‌شده")
    messages.success(request, f"وارد کردن انجام شد: {created} شاخص جدید، {updated} شاخص به‌روزرسانی‌شده، {skipped} ردیف نامعتبر رد شد.")
    return redirect("strategic:company_goals")


# ---------------- بانک شاخص‌های عملیاتی (سطح دپارتمانی، جدا از شاخص‌های کلان شرکت) ----------------

_OPKPI_EXCEL_HEADERS = [
    "کد", "عنوان شاخص", "حوزه (Q/D/C/M)", "واحد سنجش", "دپارتمان مالک", "هدف سال گذشته", "عملکرد سال گذشته",
    "هدف سال جاری", "عملکرد سال جاری", "درصد تحقق", "ترتیب نمایش",
]


def raw_factors_archive(request):
    items = list(RawIdentifiedFactor.objects.all())
    return render(request, "strategic/raw_factors_archive.html", {
        "active_page": "raw_factors_archive", "items": items, "total": len(items),
        "pestel_count": sum(1 for i in items if i.source_type == "pestel"),
        "porter_count": sum(1 for i in items if i.source_type == "porter"),
    })


def operational_kpis(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_operationalkpi" if obj_id else "strategic.add_operationalkpi"
        if _has_perm(request, perm):
            instance = get_object_or_404(OperationalKPI, pk=obj_id) if obj_id else None
            form = OperationalKPIForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE OperationalKPI" if obj_id else "CREATE OperationalKPI", str(form.instance))
                return redirect("strategic:operational_kpis")
    else:
        form = OperationalKPIForm()

    items = list(OperationalKPI.objects.all().prefetch_related("strategic_objectives__business_unit", "promoted_to_company_kpis"))

    return render(request, "strategic/operational_kpis.html", {
        "active_page": "operational_kpis", "items": items, "total": len(items),
        "form": OperationalKPIForm() if request.method == "POST" else form,
    })


@login_required
def operational_kpi_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_operationalkpi"):
        obj = get_object_or_404(OperationalKPI, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE OperationalKPI", label)
    return redirect("strategic:operational_kpis")


@login_required
def operational_kpi_export(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:operational_kpis")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "شاخص‌های عملیاتی"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(_OPKPI_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_i, k in enumerate(OperationalKPI.objects.all(), start=2):
        values = [
            k.code, k.title, k.domain, k.unit, k.department,
            k.target_1404, k.actual_1404, k.target_1405, k.actual_1405, k.progress_1405, k.order,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_i, column=col, value=val)

    widths = [12, 42, 10, 12, 28, 12, 12, 12, 12, 12, 10]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT OperationalKPI Excel", f"{OperationalKPI.objects.count()} ردیف")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="shakhes-haye-amaliati.xlsx"'
    return response


@login_required
def operational_kpi_import(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:operational_kpis")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "فایلی انتخاب نشده است.")
        return redirect("strategic:operational_kpis")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"], data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "فایل اکسل قابل خواندن نیست. لطفاً فرمت را بررسی کنید.")
        return redirect("strategic:operational_kpis")

    def _s(v):
        return "" if v is None else str(v).strip()

    created, updated, skipped = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code = _s(row[0])
        if not code:
            skipped += 1
            continue
        title = _s(row[1]) if len(row) > 1 else ""
        domain = _s(row[2]).upper() if len(row) > 2 else ""
        if domain not in ("Q", "D", "C", "M"):
            domain = "Q"
        unit = _s(row[3]) if len(row) > 3 else ""
        department = _s(row[4]) if len(row) > 4 else ""
        target_1404 = clean_number_string(_s(row[5]) if len(row) > 5 else "")
        actual_1404 = clean_number_string(_s(row[6]) if len(row) > 6 else "")
        target_1405 = clean_number_string(_s(row[7]) if len(row) > 7 else "")
        actual_1405 = clean_number_string(_s(row[8]) if len(row) > 8 else "")
        progress_1405 = clean_number_string(_s(row[9]) if len(row) > 9 else "")
        try:
            order = int(row[10]) if len(row) > 10 and row[10] not in (None, "") else 0
        except (TypeError, ValueError):
            order = 0

        _, was_created = OperationalKPI.objects.update_or_create(
            code=code,
            defaults=dict(
                title=title, domain=domain, unit=unit, department=department,
                target_1404=target_1404, actual_1404=actual_1404,
                target_1405=target_1405, actual_1405=actual_1405,
                progress_1405=progress_1405, order=order,
            ),
        )
        created += 1 if was_created else 0
        updated += 0 if was_created else 1

    _log_action(request, "IMPORT OperationalKPI Excel", f"{created} جدید، {updated} به‌روزشده، {skipped} رد‌شده")
    messages.success(request, f"وارد کردن انجام شد: {created} شاخص جدید، {updated} شاخص به‌روزرسانی‌شده، {skipped} ردیف نامعتبر رد شد.")
    return redirect("strategic:operational_kpis")


# ---------------- ورود/خروجی اکسل ماتریس اثر متقابل (فقط مدیر سیستم) ----------------

@login_required
def cross_impact_matrix_export(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:cross_impact_matrix")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    factors = list(CrossImpactFactor.objects.all().order_by("order"))
    links = {(l.from_factor_id, l.to_factor_id): l.score for l in CrossImpactLink.objects.all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ماتریس اثر متقابل"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    diag_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws.cell(row=1, column=1, value="عامل \\ عامل").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    for col, f in enumerate(factors, start=2):
        cell = ws.cell(row=1, column=col, value=f.text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_i, f in enumerate(factors, start=2):
        cell = ws.cell(row=row_i, column=1, value=f.text)
        cell.fill = header_fill
        cell.font = header_font
        for col_i, g in enumerate(factors, start=2):
            if f.pk == g.pk:
                c = ws.cell(row=row_i, column=col_i, value="")
                c.fill = diag_fill
            else:
                ws.cell(row=row_i, column=col_i, value=links.get((f.pk, g.pk), 0))

    ws.column_dimensions["A"].width = 28
    for col in range(2, len(factors) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT CrossImpactMatrix Excel", f"{len(factors)} عامل")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="matrice-asar-motaghabel.xlsx"'
    return response


@login_required
def cross_impact_matrix_import(request):
    if not request.user.is_superuser:
        messages.error(request, "این عملیات فقط برای مدیر سیستم مجاز است.")
        return redirect("strategic:cross_impact_matrix")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "فایلی انتخاب نشده است.")
        return redirect("strategic:cross_impact_matrix")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"], data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "فایل اکسل قابل خواندن نیست. لطفاً فرمت را بررسی کنید.")
        return redirect("strategic:cross_impact_matrix")

    factors_by_text = {f.text.strip(): f for f in CrossImpactFactor.objects.all()}

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_factors = []
    for cell in header_row[1:]:
        text = str(cell).strip() if cell else ""
        col_factors.append(factors_by_text.get(text))

    updated, skipped_rows, skipped_cells = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_text = str(row[0]).strip() if row and row[0] else ""
        from_factor = factors_by_text.get(row_text)
        if not from_factor:
            skipped_rows += 1
            continue
        for col_idx, to_factor in enumerate(col_factors, start=1):
            if not to_factor or to_factor.pk == from_factor.pk:
                continue
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == "":
                continue
            try:
                score = int(val)
            except (TypeError, ValueError):
                skipped_cells += 1
                continue
            if score not in (0, 1, 2, 3):
                skipped_cells += 1
                continue
            CrossImpactLink.objects.update_or_create(
                from_factor=from_factor, to_factor=to_factor, defaults={"score": score},
            )
            updated += 1

    _log_action(request, "IMPORT CrossImpactMatrix Excel", f"{updated} خونه به‌روزرسانی شد")
    msg = f"وارد کردن انجام شد: {updated} خونه به‌روزرسانی شد."
    if skipped_rows:
        msg += f" {skipped_rows} ردیف (نام عامل ناشناخته) رد شد."
    if skipped_cells:
        msg += f" {skipped_cells} خونه (مقدار نامعتبر) رد شد."
    messages.success(request, msg)
    return redirect("strategic:cross_impact_matrix")


# ---------------- ورود/خروجی اکسل کامل مخزن ذینفعان ----------------

_STAKEHOLDER_EXCEL_HEADERS = [
    "واحد/مدیریت", "نام ذینفع", "درون سازمانی (بله/خیر)", "برون سازمانی (بله/خیر)", "کانال ارتباطی",
    "نیاز/انتظار ذینفع", "نوع: نیاز (بله/خیر)", "نوع: انتظار (بله/خیر)",
    "ریسک", "احتمال وقوع ریسک", "شدت ریسک", "قابلیت تشخیص ریسک", "عدد ریسک",
    "فرصت", "امتیاز اهمیت فرصت", "امتیاز تأثیر فرصت", "عدد فرصت",
    "اقدام تعریف‌شده", "حوزه", "وضعیت رسیدگی",
]


@login_required
def stakeholder_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "مخزن ذینفعان"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(_STAKEHOLDER_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_fa = dict(Stakeholder.STATUS_CHOICES)
    for row_i, s in enumerate(Stakeholder.objects.all(), start=2):
        values = [
            s.department, s.name, "بله" if s.is_internal else "", "بله" if s.is_external else "", s.channel, s.need,
            "بله" if s.need_flag else "", "بله" if s.expectation_flag else "",
            s.risk_text, s.risk_occurrence, s.risk_severity, s.risk_detection, s.risk_score,
            s.opportunity_text, s.opportunity_importance, s.opportunity_impact, s.opportunity_score,
            s.action, s.domain, status_fa.get(s.status, s.status),
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_i, column=col, value=val)

    widths = [22, 26, 14, 14, 20, 30, 10, 10, 26, 10, 10, 10, 10, 30, 10, 10, 10, 30, 16, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT Stakeholder Excel", f"{Stakeholder.objects.count()} ردیف")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="makhzan-zinofan.xlsx"'
    return response


@login_required
def stakeholder_import(request):
    if not _has_perm(request, "strategic.add_stakeholder"):
        return redirect("strategic:stakeholders")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "فایلی انتخاب نشده است.")
        return redirect("strategic:stakeholders")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"], data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "فایل اکسل قابل خواندن نیست. لطفاً فرمت را بررسی کنید.")
        return redirect("strategic:stakeholders")

    status_by_fa = {v: k for k, v in Stakeholder.STATUS_CHOICES}

    def _s(v):
        return "" if v is None else str(v).strip()

    def _i(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    Stakeholder.objects.all().delete()
    created, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not (row[1] if len(row) > 1 else None):
            skipped += 1
            continue
        Stakeholder.objects.create(
            department=_s(row[0]),
            name=_s(row[1]),
            is_internal=_s(row[2]).startswith("بل") if len(row) > 2 else False,
            is_external=_s(row[3]).startswith("بل") if len(row) > 3 else False,
            channel=_s(row[4]) if len(row) > 4 else "",
            need=_s(row[5]) if len(row) > 5 else "",
            need_flag=_s(row[6]).startswith("بل") if len(row) > 6 else False,
            expectation_flag=_s(row[7]).startswith("بل") if len(row) > 7 else False,
            risk_text=_s(row[8]) if len(row) > 8 else "",
            risk_occurrence=_i(row[9]) if len(row) > 9 else None,
            risk_severity=_i(row[10]) if len(row) > 10 else None,
            risk_detection=_i(row[11]) if len(row) > 11 else None,
            risk_score=_i(row[12]) if len(row) > 12 else None,
            opportunity_text=_s(row[13]) if len(row) > 13 else "",
            opportunity_importance=_i(row[14]) if len(row) > 14 else None,
            opportunity_impact=_i(row[15]) if len(row) > 15 else None,
            opportunity_score=_i(row[16]) if len(row) > 16 else None,
            action=_s(row[17]) if len(row) > 17 else "",
            domain=_s(row[18]) if len(row) > 18 else "",
            status=status_by_fa.get(_s(row[19]), "open") if len(row) > 19 else "open",
        )
        created += 1

    _log_action(request, "IMPORT Stakeholder Excel (replace-all)", f"{created} ردیف جدید، {skipped} رد‌شده")
    messages.success(request, f"جایگزینی انجام شد: مخزن قبلی پاک شد و {created} ذینفع از فایل جدید ثبت شد. {skipped} ردیف نامعتبر رد شد.")
    return redirect("strategic:stakeholders")


# ---------------- بانک الزامات قانونی ----------------

def legal_requirements(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_legalrequirement" if obj_id else "strategic.add_legalrequirement"
        if _has_perm(request, perm):
            instance = get_object_or_404(LegalRequirement, pk=obj_id) if obj_id else None
            form = LegalRequirementForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE LegalRequirement" if obj_id else "CREATE LegalRequirement", str(form.instance))
                return redirect("strategic:legal_requirements")
        else:
            form = LegalRequirementForm()
    else:
        form = LegalRequirementForm()

    items = LegalRequirement.objects.all().select_related("related_pestel")
    q = request.GET.get("q", "").strip()
    dept = request.GET.get("dept", "").strip()
    if q:
        items = items.filter(
            Q(title__icontains=q) | Q(source__icontains=q) | Q(risk_text__icontains=q) | Q(opportunity_text__icontains=q)
        )
    if dept:
        items = items.filter(department=dept)

    all_items = list(LegalRequirement.objects.all())
    departments = sorted({i.department for i in all_items if i.department})

    total = len(all_items)
    legal_n = sum(1 for i in all_items if i.is_legal)
    org_n = sum(1 for i in all_items if i.is_organizational)
    internal_n = sum(1 for i in all_items if i.is_internal)
    external_n = sum(1 for i in all_items if i.is_external)
    with_risk_n = sum(1 for i in all_items if i.risk_text)
    with_opp_n = sum(1 for i in all_items if i.opportunity_text)
    linked_pestel_n = sum(1 for i in all_items if i.related_pestel_id)

    def pct(part, whole):
        return round(part / whole * 100) if whole else 0

    lo_sum = max(legal_n + org_n, 1)
    ie_sum = max(internal_n + external_n, 1)

    summary = {
        "total": total,
        "legal": legal_n, "organizational": org_n,
        "legal_pct": round(legal_n / lo_sum * 100), "org_pct": round(org_n / lo_sum * 100),
        "internal": internal_n, "external": external_n,
        "internal_pct": round(internal_n / ie_sum * 100), "external_pct": round(external_n / ie_sum * 100),
        "with_risk": with_risk_n, "with_risk_pct": pct(with_risk_n, total),
        "with_opportunity": with_opp_n, "with_opportunity_pct": pct(with_opp_n, total),
        "linked_pestel": linked_pestel_n, "linked_pestel_pct": pct(linked_pestel_n, total),
    }

    return render(request, "strategic/legal_requirements.html", {
        "active_page": "legal_requirements", "items": items, "form": form, "q": q, "dept": dept,
        "departments": departments, "total_count": total, "summary": summary,
    })


@login_required
def legal_requirement_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_legalrequirement"):
        obj = get_object_or_404(LegalRequirement, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE LegalRequirement", label)
    return redirect("strategic:legal_requirements")


_LEGAL_EXCEL_HEADERS = [
    "الزامات قانونی و سازمانی", "مأخذ الزام", "قانونی (بله/خیر)", "سازمانی (بله/خیر)",
    "درون سازمانی (بله/خیر)", "برون سازمانی (بله/خیر)", "تاریخ ویرایش الزام",
    "مستندات داخلی مرتبط", "محل کاربرد", "ریسک", "فرصت", "توضیحات", "نام مدیریت/معاونت",
]


@login_required
def legal_requirement_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "بانک الزامات قانونی"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(_LEGAL_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_i, r in enumerate(LegalRequirement.objects.all(), start=2):
        values = [
            r.title, r.source, "بله" if r.is_legal else "", "بله" if r.is_organizational else "",
            "بله" if r.is_internal else "", "بله" if r.is_external else "",
            r.revision_date, r.related_documents, r.scope, r.risk_text, r.opportunity_text,
            r.notes, r.department,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_i, column=col, value=val)

    widths = [34, 20, 12, 12, 14, 14, 16, 28, 20, 28, 28, 24, 22]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT LegalRequirement Excel", f"{LegalRequirement.objects.count()} ردیف")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bank-elzamat-ghanooni.xlsx"'
    return response


@login_required
def legal_requirement_import(request):
    if not _has_perm(request, "strategic.add_legalrequirement"):
        return redirect("strategic:legal_requirements")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "فایلی انتخاب نشده است.")
        return redirect("strategic:legal_requirements")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"], data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "فایل اکسل قابل خواندن نیست. لطفاً فرمت را بررسی کنید.")
        return redirect("strategic:legal_requirements")

    def _s(v):
        v = "" if v is None else str(v).strip()
        return "" if v == "-" else v

    LegalRequirement.objects.all().delete()
    created, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not (row[0] if len(row) > 0 else None):
            skipped += 1
            continue
        LegalRequirement.objects.create(
            title=_s(row[0]),
            source=_s(row[1]) if len(row) > 1 else "",
            is_legal=_s(row[2]).startswith("بل") if len(row) > 2 else False,
            is_organizational=_s(row[3]).startswith("بل") if len(row) > 3 else False,
            is_internal=_s(row[4]).startswith("بل") if len(row) > 4 else False,
            is_external=_s(row[5]).startswith("بل") if len(row) > 5 else False,
            revision_date=_s(row[6]) if len(row) > 6 else "",
            related_documents=_s(row[7]) if len(row) > 7 else "",
            scope=_s(row[8]) if len(row) > 8 else "",
            risk_text=_s(row[9]) if len(row) > 9 else "",
            opportunity_text=_s(row[10]) if len(row) > 10 else "",
            notes=_s(row[11]) if len(row) > 11 else "",
            department=_s(row[12]) if len(row) > 12 else "",
        )
        created += 1

    _log_action(request, "IMPORT LegalRequirement Excel (replace-all)", f"{created} ردیف جدید، {skipped} رد‌شده")
    messages.success(request, f"جایگزینی انجام شد: بانک قبلی پاک شد و {created} الزام از فایل جدید ثبت شد. {skipped} ردیف نامعتبر رد شد.")
    return redirect("strategic:legal_requirements")


# ---------------- بانک عوامل محیطی ----------------

def environmental_factors(request):
    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        perm = "strategic.change_environmentalfactor" if obj_id else "strategic.add_environmentalfactor"
        if _has_perm(request, perm):
            instance = get_object_or_404(EnvironmentalFactor, pk=obj_id) if obj_id else None
            form = EnvironmentalFactorForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                _log_action(request, "UPDATE EnvironmentalFactor" if obj_id else "CREATE EnvironmentalFactor", str(form.instance))
                return redirect("strategic:environmental_factors")
        else:
            form = EnvironmentalFactorForm()
    else:
        form = EnvironmentalFactorForm()

    items = EnvironmentalFactor.objects.all()
    q = request.GET.get("q", "").strip()
    cat = request.GET.get("cat", "").strip()
    if q:
        items = items.filter(
            Q(factor_text__icontains=q) | Q(category__icontains=q) | Q(detail__icontains=q) | Q(effect_type__icontains=q)
        )
    if cat:
        items = items.filter(category=cat)

    all_items = list(EnvironmentalFactor.objects.all())
    categories = sorted({i.category for i in all_items if i.category})

    category_order = {
        "عوامل سیاسی": 1, "عوامل اقتصادی": 2, "عوامل اجتماعی": 3,
        "عوامل فناوری و تکنولوژی": 4, "عوامل زیست محیطی": 5,
        "قوانین و الزامات": 6, "عوامل قوانین و الزامات": 6,
        "قدرت رقبای موجود": 11, "قدرت چانه زنی خریداران": 12,
        "قدرت چانه زنی تامین کنندگان": 13, "تهدید تازه واردان به صنعت بازرگانی قطعات و ارائه خدمات": 14,
        "تهدید قطعات و خدمات جایگزین": 15,
    }
    items = sorted(items, key=lambda i: (category_order.get(i.category, 99), -(i.avg_score or 0), i.order))

    return render(request, "strategic/environmental_factors.html", {
        "active_page": "environmental_factors", "items": items, "form": form, "q": q, "cat": cat,
        "categories": categories, "total_count": len(all_items),
    })


@login_required
def environmental_factor_delete(request, pk):
    if request.method == "POST" and _has_perm(request, "strategic.delete_environmentalfactor"):
        obj = get_object_or_404(EnvironmentalFactor, pk=pk)
        label = str(obj)
        obj.delete()
        _log_action(request, "DELETE EnvironmentalFactor", label)
    return redirect("strategic:environmental_factors")


_ENV_FACTOR_EXCEL_HEADERS = [
    "ردیف", "دسته‌بندی محیط", "شرح عامل تأثیرگذار", "توضیح تفصیلی", "نوع اثر", "راهنمای امتیازدهی",
    "میانگین امتیاز", "فراوانی اثر بالا (۷-۸)", "فراوانی اثر بسیار بالا (۹-۱۰)", "جمع فراوانی اثرهای بالا",
]


@login_required
def environmental_factor_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "بانک عوامل محیطی"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(_ENV_FACTOR_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_i, r in enumerate(EnvironmentalFactor.objects.all().order_by("-avg_score", "order"), start=2):
        values = [
            r.order, r.category, r.factor_text, r.detail, r.effect_type, r.scoring_guide,
            float(r.avg_score) if r.avg_score is not None else None,
            r.freq_high, r.freq_very_high, r.freq_total,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_i, column=col, value=val)

    widths = [8, 22, 34, 40, 14, 26, 12, 12, 14, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT EnvironmentalFactor Excel", f"{EnvironmentalFactor.objects.count()} ردیف")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bank-avamel-mohiti.xlsx"'
    return response


@login_required
def environmental_factor_import(request):
    if not _has_perm(request, "strategic.add_environmentalfactor"):
        return redirect("strategic:environmental_factors")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "فایلی انتخاب نشده است.")
        return redirect("strategic:environmental_factors")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"], data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "فایل اکسل قابل خواندن نیست. لطفاً فرمت را بررسی کنید.")
        return redirect("strategic:environmental_factors")

    def _s(v):
        return "" if v is None else str(v).strip()

    def _i(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    EnvironmentalFactor.objects.all().delete()
    created, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not (row[2] if len(row) > 2 else None):
            skipped += 1
            continue
        EnvironmentalFactor.objects.create(
            order=_i(row[0]) or 0,
            category=_s(row[1]) if len(row) > 1 else "",
            factor_text=_s(row[2]),
            detail=_s(row[3]) if len(row) > 3 else "",
            effect_type=_s(row[4]) if len(row) > 4 else "",
            scoring_guide=_s(row[5]) if len(row) > 5 else "",
            avg_score=_f(row[6]) if len(row) > 6 else None,
            freq_high=_i(row[7]) if len(row) > 7 else None,
            freq_very_high=_i(row[8]) if len(row) > 8 else None,
            freq_total=_i(row[9]) if len(row) > 9 else None,
        )
        created += 1

    _log_action(request, "IMPORT EnvironmentalFactor Excel (replace-all)", f"{created} ردیف جدید، {skipped} رد‌شده")
    messages.success(request, f"جایگزینی انجام شد: بانک قبلی پاک شد و {created} عامل از فایل جدید ثبت شد. {skipped} ردیف نامعتبر رد شد.")
    return redirect("strategic:environmental_factors")


# ---------------- خروجی اکسل جامع نقشه استراتژیک (۴ شیت، هر کسب‌وکار جدا) ----------------

def _kpi_pct_color(pct):
    if pct is None:
        return None
    if pct < 60:
        return "red"
    if pct < 90:
        return "yellow"
    return "green"


def _objective_kpi_entries(o):
    """لیست شاخص‌های وصل به یک هدف (مشترک + اختصاصی) با کد/نام/هدف/عملکرد/درصد."""
    entries = []
    for k in o.linked_kpis.all():
        entries.append({
            "code": k.code, "name": k.name, "target": k.target_1405, "actual": k.actual_1405,
            "pct": k.manual_progress_value,
        })
    for k in o.kpis.all():
        entries.append({
            "code": "", "name": k.name, "target": k.target, "actual": k.actual,
            "pct": k.progress_pct,
        })
    return entries


def _objective_overall_pct(entries):
    vals = [e["pct"] for e in entries if e["pct"] is not None]
    return round(sum(vals) / len(vals)) if vals else None


def _objective_swot_items(o):
    return list(o.source_tows.source_items.all()) if o.source_tows_id else []


@login_required
def stratmap_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter

    business_units = list(BusinessUnit.objects.all())
    bu_id = request.GET.get("bu")
    current_bu = None
    if bu_id:
        current_bu = next((b for b in business_units if str(b.pk) == str(bu_id)), None)
    if not current_bu and business_units:
        current_bu = business_units[0]
    if not current_bu:
        messages.error(request, "هیچ کسب‌وکاری برای خروجی وجود ندارد.")
        return redirect("strategic:stratmap")

    objectives = list(
        StrategicObjective.objects.filter(business_unit=current_bu)
        .select_related("theme", "source_tows")
        .prefetch_related("feeds_into", "fed_by", "linked_kpis", "kpis", "source_tows__source_items")
    )

    HEADER_FILL = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10.5)
    WRAP = Alignment(horizontal="right", vertical="top", wrap_text=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    GREEN_FILL = PatternFill(start_color="D9EAD9", end_color="D9EAD9", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid")
    RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    PCT_FILL = {"green": GREEN_FILL, "yellow": YELLOW_FILL, "red": RED_FILL}
    STATUS_EMOJI = {"green": "🟢 مطلوب", "yellow": "🟡 نیازمند توجه", "red": "🔴 بحرانی"}

    def style_header_row(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    wb = openpyxl.Workbook()

    # ============= شیت ۱: داشبورد استراتژیک =============
    ws1 = wb.active
    ws1.title = "داشبورد استراتژیک"
    ws1.sheet_view.rightToLeft = True
    ws1.cell(row=1, column=1, value=f"داشبورد استراتژیک — {current_bu.name}").font = Font(bold=True, size=14)
    ws1.merge_cells("A1:F1")

    headers1 = ["منظر", "تعداد اهداف", "تحقق میانگین", "سبز", "زرد", "قرمز"]
    for c, h in enumerate(headers1, start=1):
        ws1.cell(row=3, column=c, value=h)
    style_header_row(ws1, 3, len(headers1))

    persp_map = [("financial", "مالی"), ("customer", "مشتری"), ("process", "فرآیند"), ("learning", "یادگیری و رشد")]
    row = 4
    for p_key, p_label in persp_map:
        p_objs = [o for o in objectives if o.perspective == p_key]
        pcts = []
        g = y = r = 0
        for o in p_objs:
            entries = _objective_kpi_entries(o)
            pct = _objective_overall_pct(entries)
            if pct is not None:
                pcts.append(pct)
                color = _kpi_pct_color(pct)
                if color == "green":
                    g += 1
                elif color == "yellow":
                    y += 1
                else:
                    r += 1
        avg_pct = round(sum(pcts) / len(pcts)) if pcts else None
        ws1.cell(row=row, column=1, value=p_label)
        ws1.cell(row=row, column=2, value=len(p_objs))
        ws1.cell(row=row, column=3, value=(f"{avg_pct}%" if avg_pct is not None else "—"))
        ws1.cell(row=row, column=4, value=g)
        ws1.cell(row=row, column=5, value=y)
        ws1.cell(row=row, column=6, value=r)
        for c in range(1, 7):
            ws1.cell(row=row, column=c).border = BORDER
            ws1.cell(row=row, column=c).alignment = CENTER
        row += 1

    row += 2
    ws1.cell(row=row, column=1, value="خلاصه کلی").font = Font(bold=True, size=12)
    row += 1
    summary_rows = [
        ("تعداد اهداف استراتژیک", len(objectives)),
        ("تعداد راهبردهای TOWS", sum(1 for o in objectives if o.source_tows_id)),
        ("تعداد شاخص‌های وصل‌شده (مشترک+اختصاصی)", sum(len(_objective_kpi_entries(o)) for o in objectives)),
        ("تعداد پروژه‌های تحول", Initiative.objects.filter(business_unit=current_bu).count()),
        ("تعداد ریسک‌های وصل به این کسب‌وکار", Risk.objects.filter(linked_objective__business_unit=current_bu).distinct().count()),
    ]
    for label, val in summary_rows:
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=val)
        row += 1

    for col, w in zip("ABCDEF", [22, 14, 14, 8, 8, 8]):
        ws1.column_dimensions[col].width = w

    # ============= شیت ۲: نقشه استراتژیک =============
    ws2 = wb.create_sheet("نقشه استراتژیک")
    ws2.sheet_view.rightToLeft = True
    headers2 = ["کد", "هدف استراتژیک", "منظر", "محور", "KPIها", "تحقق", "وضعیت", "TOWS", "SWOT", "اهداف مرتبط", "منشأ تحلیل"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=c, value=h)
    style_header_row(ws2, 1, len(headers2))
    ws2.freeze_panes = "A2"

    PERSP_LABEL = dict(StrategicObjective.PERSPECTIVE_CHOICES)
    r = 2
    for o in objectives:
        entries = _objective_kpi_entries(o)
        overall_pct = _objective_overall_pct(entries)
        color = _kpi_pct_color(overall_pct)

        kpi_lines = []
        for e in entries:
            label = f"{e['code']} {e['name']}".strip() if e["code"] else e["name"]
            pct_txt = f"{e['pct']}%" if e["pct"] is not None else "—"
            kpi_lines.append(f"{label}\nهدف:{e['target'] or '—'}  عملکرد:{e['actual'] or '—'}  تحقق:{pct_txt}")
        kpi_cell = "\n\n".join(kpi_lines) if kpi_lines else "—"

        tows_cell = f"{o.source_tows.get_category_display()}\n{o.source_tows.text}" if o.source_tows_id else "—"

        swot_items = _objective_swot_items(o)
        swot_groups = {}
        for it in swot_items:
            swot_groups.setdefault(it.category, []).append(it)
        swot_lines = []
        cat_label = {"s": "قوت‌ها", "w": "ضعف‌ها", "o": "فرصت‌ها", "t": "تهدیدها"}
        for cat in ["s", "o", "w", "t"]:
            if cat in swot_groups:
                for it in swot_groups[cat]:
                    swot_lines.append(f"{it.category.upper()}: {it.text}")
        swot_cell = "\n".join(swot_lines) if swot_lines else "—"

        conn_lines = []
        for t in o.feeds_into.all():
            conn_lines.append(f"→ {t.code}\n{t.title}")
        for f in o.fed_by.all():
            conn_lines.append(f"← {f.code}\n{f.title}")
        conn_cell = "\n\n".join(conn_lines) if conn_lines else "—"

        source_lines = []
        for it in swot_items:
            if it.source_label:
                source_lines.append(it.source_label)
        source_cell = "\n".join(source_lines) if source_lines else "—"

        ws2.cell(row=r, column=1, value=o.code)
        ws2.cell(row=r, column=2, value=o.title)
        ws2.cell(row=r, column=3, value=PERSP_LABEL.get(o.perspective, o.perspective))
        ws2.cell(row=r, column=4, value=o.theme.name if o.theme_id else "—")
        ws2.cell(row=r, column=5, value=kpi_cell)
        pct_cell = ws2.cell(row=r, column=6, value=(overall_pct / 100 if overall_pct is not None else None))
        pct_cell.number_format = "0%"
        ws2.cell(row=r, column=7, value=STATUS_EMOJI.get(color, "—"))
        ws2.cell(row=r, column=8, value=tows_cell)
        ws2.cell(row=r, column=9, value=swot_cell)
        ws2.cell(row=r, column=10, value=conn_cell)
        ws2.cell(row=r, column=11, value=source_cell)

        for c in range(1, 12):
            cell = ws2.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
        if color:
            ws2.cell(row=r, column=7).fill = PCT_FILL[color]
        r += 1

    widths2 = [8, 26, 10, 14, 34, 9, 15, 26, 30, 26, 26]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for rr in range(2, r):
        ws2.row_dimensions[rr].height = 70

    if r > 2:
        databar = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="3E7A52")
        ws2.conditional_formatting.add(f"F2:F{r-1}", databar)

    # ============= شیت ۳: Strategic Traceability =============
    ws3 = wb.create_sheet("Strategic Traceability")
    ws3.sheet_view.rightToLeft = True
    headers3 = ["کد", "هدف", "TOWS", "SWOT", "منشأ SWOT", "KPI"]
    for c, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=c, value=h)
    style_header_row(ws3, 1, len(headers3))
    ws3.freeze_panes = "A2"

    r = 2
    for o in objectives:
        tows_cell = f"{o.source_tows.get_category_display()}\n{o.source_tows.text[:120]}" if o.source_tows_id else "—"
        swot_items = _objective_swot_items(o)
        swot_cell = "\n".join(f"{it.category.upper()}: {it.text}" for it in swot_items) if swot_items else "—"
        source_cell = "\n".join(it.source_label for it in swot_items if it.source_label) if swot_items else "—"
        entries = _objective_kpi_entries(o)
        kpi_cell = "\n".join(
            f"{(e['code'] + ' ' if e['code'] else '')}{e['name']} — {e['pct']}%" if e["pct"] is not None
            else f"{(e['code'] + ' ' if e['code'] else '')}{e['name']}"
            for e in entries
        ) if entries else "—"

        ws3.cell(row=r, column=1, value=o.code)
        ws3.cell(row=r, column=2, value=o.title)
        ws3.cell(row=r, column=3, value=tows_cell)
        ws3.cell(row=r, column=4, value=swot_cell)
        ws3.cell(row=r, column=5, value=source_cell)
        ws3.cell(row=r, column=6, value=kpi_cell)
        for c in range(1, 7):
            cell = ws3.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1

    widths3 = [8, 26, 30, 30, 26, 26]
    for i, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for rr in range(2, r):
        ws3.row_dimensions[rr].height = 80

    # ============= شیت ۴: همه KPIهای مرتبط با این کسب‌وکار =============
    ws4 = wb.create_sheet("همه KPIها")
    ws4.sheet_view.rightToLeft = True
    headers4 = ["کد", "شاخص", "نوع", "هدف", "عملکرد", "تحقق", "اهداف مرتبط", "SWOT مرتبط (از طریق اهداف)"]
    for c, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=c, value=h)
    style_header_row(ws4, 1, len(headers4))
    ws4.freeze_panes = "A2"

    # جمع‌آوری منحصربه‌فرد شاخص‌های مشترک و اختصاصی مرتبط با اهداف این کسب‌وکار
    shared_map = {}   # kpi_pk -> {"kpi": obj, "objectives": [o,...]}
    custom_list = []  # list of (StrategicKPI, objective)
    for o in objectives:
        for k in o.linked_kpis.all():
            shared_map.setdefault(k.pk, {"kpi": k, "objectives": []})
            shared_map[k.pk]["objectives"].append(o)
        for k in o.kpis.all():
            custom_list.append((k, o))

    r = 2
    for entry in shared_map.values():
        k = entry["kpi"]
        objs = entry["objectives"]
        pct = k.manual_progress_value
        color = _kpi_pct_color(pct)
        obj_cell = "\n".join(f"{o.code} — {o.title}" for o in objs)
        swot_lines = []
        for o in objs:
            for it in _objective_swot_items(o):
                swot_lines.append(f"{it.category.upper()}: {it.text}")
        swot_cell = "\n".join(dict.fromkeys(swot_lines)) if swot_lines else "—"

        ws4.cell(row=r, column=1, value=k.code)
        ws4.cell(row=r, column=2, value=k.name)
        ws4.cell(row=r, column=3, value="مشترک شرکت")
        ws4.cell(row=r, column=4, value=k.target_1405 or "—")
        ws4.cell(row=r, column=5, value=k.actual_1405 or "—")
        ws4.cell(row=r, column=6, value=(f"{pct}%" if pct is not None else "—"))
        ws4.cell(row=r, column=7, value=obj_cell)
        ws4.cell(row=r, column=8, value=swot_cell)
        if color:
            ws4.cell(row=r, column=6).fill = PCT_FILL[color]
        for c in range(1, 9):
            cell = ws4.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1

    for k, o in custom_list:
        pct = k.progress_pct
        color = _kpi_pct_color(pct)
        swot_cell = "\n".join(f"{it.category.upper()}: {it.text}" for it in _objective_swot_items(o)) or "—"

        ws4.cell(row=r, column=1, value="—")
        ws4.cell(row=r, column=2, value=k.name)
        ws4.cell(row=r, column=3, value="اختصاصی هدف")
        ws4.cell(row=r, column=4, value=k.target or "—")
        ws4.cell(row=r, column=5, value=k.actual or "—")
        ws4.cell(row=r, column=6, value=(f"{pct}%" if pct is not None else "—"))
        ws4.cell(row=r, column=7, value=f"{o.code} — {o.title}")
        ws4.cell(row=r, column=8, value=swot_cell)
        if color:
            ws4.cell(row=r, column=6).fill = PCT_FILL[color]
        for c in range(1, 9):
            cell = ws4.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1

    widths4 = [8, 28, 14, 12, 12, 10, 30, 30]
    for i, w in enumerate(widths4, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    for rr in range(2, r):
        ws4.row_dimensions[rr].height = 46

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _log_action(request, "EXPORT StratMap Excel (4 sheets)", f"{current_bu.name} — {len(objectives)} هدف")
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    safe_name = current_bu.name.replace(" ", "-")
    response["Content-Disposition"] = f'attachment; filename="nagshe-esteratejik-{safe_name}.xlsx"'
    return response


# ============================== مدیریت کاربران ==============================

def _superuser_required(request):
    if not request.user.is_authenticated or not request.user.is_superuser:
        return False
    return True


@login_required
def user_list(request):
    if not _superuser_required(request):
        messages.error(request, "دسترسی به این بخش فقط برای مدیر سامانه مجاز است.")
        return redirect("strategic:home")

    from django.contrib.auth.models import User
    users = User.objects.all().select_related("profile").order_by("-date_joined")
    return render(request, "strategic/user_list.html", {
        "active_page": "user_list", "users": users,
    })


@login_required
def user_edit(request, pk=None):
    if not _superuser_required(request):
        messages.error(request, "دسترسی به این بخش فقط برای مدیر سامانه مجاز است.")
        return redirect("strategic:home")

    from django.contrib.auth.models import User, Group
    from strategic.models import UserProfile, BusinessUnit

    instance = get_object_or_404(User, pk=pk) if pk else None

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip()
        role = request.POST.get("role", "viewer")
        bu_id = request.POST.get("business_unit") or None
        phone = request.POST.get("phone", "").strip()
        new_password = request.POST.get("new_password", "").strip()
        is_active = request.POST.get("is_active") == "on"

        if not username:
            messages.error(request, "نام کاربری الزامی است.")
            return redirect("strategic:user_edit", pk=pk) if pk else redirect("strategic:user_add")

        if instance:
            user_obj = instance
            if User.objects.exclude(pk=user_obj.pk).filter(username=username).exists():
                messages.error(request, "این نام کاربری قبلاً استفاده شده است.")
                return redirect("strategic:user_edit", pk=pk)
            user_obj.username = username
        else:
            if User.objects.filter(username=username).exists():
                messages.error(request, "این نام کاربری قبلاً استفاده شده است.")
                return redirect("strategic:user_add")
            if not new_password:
                messages.error(request, "برای کاربر جدید، تعیین رمز عبور الزامی است.")
                return redirect("strategic:user_add")
            user_obj = User(username=username)

        user_obj.first_name = first_name
        user_obj.last_name = last_name
        user_obj.email = email
        user_obj.is_active = is_active
        user_obj.is_staff = True  # برای دسترسی به بخش‌های مدیریتی سامانه لازم است
        user_obj.is_superuser = (role == "admin")
        if new_password:
            user_obj.set_password(new_password)
        user_obj.save()

        # اعمال گروه بر اساس نقش
        user_obj.groups.clear()
        if role == "editor":
            editor_group, _ = Group.objects.get_or_create(name="ویرایشگر")
            user_obj.groups.add(editor_group)
        elif role == "viewer":
            viewer_group, _ = Group.objects.get_or_create(name="کارشناس (فقط مشاهده)")
            user_obj.groups.add(viewer_group)

        profile, _ = UserProfile.objects.get_or_create(user=user_obj)
        profile.role = role
        profile.phone = phone
        profile.business_unit_id = bu_id
        profile.save()

        action = "UPDATE User" if instance else "CREATE User"
        _log_action(request, action, f"{user_obj.username} — نقش: {role}")
        messages.success(request, f"کاربر «{username}» با موفقیت ذخیره شد.")
        return redirect("strategic:user_list")

    profile = getattr(instance, "profile", None) if instance else None
    return render(request, "strategic/user_form.html", {
        "active_page": "user_list", "instance": instance, "profile": profile,
        "business_units": BusinessUnit.objects.all(),
        "role_choices": UserProfile.ROLE_CHOICES,
    })


@login_required
def user_toggle_active(request, pk):
    if not _superuser_required(request) or request.method != "POST":
        messages.error(request, "این عملیات مجاز نیست.")
        return redirect("strategic:user_list")

    from django.contrib.auth.models import User
    user_obj = get_object_or_404(User, pk=pk)
    if user_obj.pk == request.user.pk:
        messages.error(request, "نمی‌توانید حساب کاربری خودتان را غیرفعال کنید.")
        return redirect("strategic:user_list")

    user_obj.is_active = not user_obj.is_active
    user_obj.save(update_fields=["is_active"])
    _log_action(request, "TOGGLE User active", f"{user_obj.username} -> {user_obj.is_active}")
    messages.success(request, f"وضعیت کاربر «{user_obj.username}» تغییر کرد.")
    return redirect("strategic:user_list")
